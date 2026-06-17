import undetected_chromedriver as uc
import re
import pandas as pd
import sys
import time
import os
import random
import openpyxl
import logging
import shutil
import traceback
import math
import calendar
import holidays
import pyautogui
import selenium_stealth
import pywin
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from pathlib import Path
from utils import (
    carregar_dados_planilha_mensal,
    ajustar_largura_colunas,
    CAMINHO_PLANILHA,
    PASTA_DOWNLOADS,
    PASTA_BASE,
    MESES_ABREV,
    MESES_NUMERO,
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)

def configurar_navegador_com_certificado(caminho_certificado, nome_certificado):
    pasta_download_base = PASTA_DOWNLOADS
    subpasta = "Parcelamentos ADM - SP"
    pasta_download = criar_estrutura_pastas_por_vencimento(pasta_download_base, subpasta)

    logging.info(f"Usando pasta de download: {pasta_download}")
    
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["disable-popup-blocking"])
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")
    options.add_argument(f"--ssl-client-certificate={caminho_certificado}")
    options.add_argument(f"--ssl-client-key={caminho_certificado}")
    
    prefs = {
        "download.default_directory": pasta_download,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "AutoSelectCertificateForUrls": [{
            "pattern": "identityprd.fazenda.sp.gov.br",
            "filter":{
                "SUBJECT": {"CN": nome_certificado}
            }
        }]
    }
    options.add_experimental_option("prefs", prefs)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    selenium_stealth.stealth(
        driver,
        languages=["pt-BR", "pt"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
    )
    
    return driver, pasta_download

def selecionar_certificado(nome_arquivo_img=None):
    """Seleciona certificado com múltiplas estratégias"""
    
    logging.info("=== INICIANDO SELEÇÃO DE CERTIFICADO ===")
    time.sleep(2)
    
    # Estratégia 1: PyWinAuto (mais robusta)
    try:
        if selecionar_certificado_pywinauto():
            logging.info("✅ Certificado selecionado via PyWinAuto")
            return True
    except Exception as e:
        logging.warning(f"PyWinAuto falhou: {e}")
    
    # Estratégia 2: Navegação por teclado
    try:
        if selecionar_certificado_teclado():
            logging.info("✅ Certificado selecionado via teclado")
            return True
    except Exception as e:
        logging.warning(f"Navegação por teclado falhou: {e}")
    
    # Estratégia 3: Fallback manual
    logging.warning("⚠️ Automação falhou - solicitando seleção manual")
    print("\n" + "="*60)
    print("SELECIONE O CERTIFICADO MANUALMENTE:")
    print("1. Na janela de certificados que abriu")
    print("2. Procure pelo Certificado")
    print("3. Clique no certificado e depois em OK")
    print("4. Volte aqui e pressione ENTER")
    print("="*60)
    input("Pressione ENTER após selecionar o certificado...")
    
    return True

def selecionar_certificado_pywinauto():
    try:
        from pywinauto import Application
        from pywinauto.findwindows import find_windows
        import time
        
        logging.info("Aguardando janela de certificados...")
        time.sleep(2)
        
        # Encontrar a janela de seleção de certificados
        windows = find_windows(title_re=".*[Ss]elecionar.*[Cc]ertificado.*")
        if not windows:
            windows = find_windows(title_re=".*[Cc]ertificate.*")
        
        if windows:
            app = Application().connect(handle=windows[0])
            dialog = app.window(handle=windows[0])
            
            # Procurar por lista de certificados
            cert_list = dialog.child_window(class_name="ListBox")
            if not cert_list.exists():
                cert_list = dialog.child_window(class_name="ListView")
            
            if cert_list.exists():
                # Navegar pela lista procurando
                items = cert_list.get_items()
                for i, item in enumerate(items):
                    item_text = str(item)
                    if "Certificado" in item_text.upper():
                        cert_list.select(i)
                        logging.info(f"Certificado encontrado: {item_text}")
                        
                        # Clicar OK
                        ok_button = dialog.child_window(title="OK")
                        if ok_button.exists():
                            ok_button.click()
                        else:
                            pyautogui.press("enter")
                        return True
        
        logging.warning("Certificado não encontrado")
        return False
        
    except Exception as e:
        logging.error(f"Erro no PyWinAuto: {e}")
        return False

def selecionar_certificado_teclado():
    logging.info("Navegando pelos certificados com teclado...")
    time.sleep(2)
    
    # Ir para o primeiro certificado
    pyautogui.press('home')
    time.sleep(0.5)
    
    # Navegar até encontrar(máximo 150 tentativas)
    for i in range(150):
        # Copiar texto selecionado para clipboard
        pyautogui.hotkey('ctrl', 'c')
        time.sleep(0.2)
        
        try:
            import clipboard
            texto_certificado = clipboard.paste().upper()
            
            if "Certificado" in texto_certificado:
                logging.info(f"Certificado encontrado na posição {i}")
                pyautogui.press('enter')
                return True
                
        except:
            pass
        
        # Próximo certificado
        pyautogui.press('down')
        time.sleep(0.1)
    
    logging.warning("Certificado não encontrado após 150 tentativas")
    return False
    pass

def limpar_nome_arquivo(nome_empresa):
    caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']

    nome_limpo = nome_empresa
    for char in caracteres_invalidos:
        nome_limpo = nome_limpo.replace(char, '_')
    
    if len(nome_limpo) > 50:
        nome_limpo = nome_limpo[:47] + '...'
    
    return nome_limpo.strip()

def limpar_cnpj(cnpj_sujo):
    if pd.isna(cnpj_sujo) or cnpj_sujo == '':
        return ''
    
    try:
        cnpj_str = str(int(float(cnpj_sujo)))
    
    except:
        cnpj_str = str(cnpj_sujo)
    
    apenas_numeros = ''.join(char for char in cnpj_str if char.isdigit())

    return apenas_numeros.zfill(14)

def criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, data_vencimento=None):
    meses_abrev = {
        1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR',
        5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
        9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
    }
    
    if data_vencimento:
        try:
            partes_data = data_vencimento.split('/')
            
            mes_raw = partes_data[1].strip().upper()
            ano = partes_data[2].strip()
            
            if mes_raw.isdigit():
                mes_num = int(mes_raw)

            else:
                meses_str_para_num = {
                    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4,
                    'MAI': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
                    'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
                }
                mes_abrev = mes_raw[:3]
                mes_num = meses_str_para_num.get(mes_abrev, datetime.now().month)
            
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
            
        except Exception as e:
            logging.warning(f"Erro ao processar data de vencimento '{data_vencimento}', usando mês atual: {str(e)}")
            data_atual = datetime.now()
            mes_num = data_atual.month
            ano = str(data_atual.year)
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"

    else:
        data_atual = datetime.now()
        mes_num = data_atual.month
        ano = str(data_atual.year)
        mes_abrev_pasta = meses_abrev[mes_num]
        meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
        numero_mes = meses_dict[mes_abrev_pasta.upper()]
        nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
    
    caminho_pasta_mes = os.path.join(pasta_base, nome_pasta_mes)
    
    if not os.path.exists(caminho_pasta_mes):
        os.makedirs(caminho_pasta_mes)
        logging.info(f"Pasta do mês criada: {caminho_pasta_mes}")
    
    caminho_completo = os.path.join(caminho_pasta_mes, subpasta)
    
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
        logging.info(f"Subpasta criada: {caminho_completo}")
    
    return caminho_completo

def baixar_dae_mes_atual(driver, num_parcelamento, mes_atual_texto, ano_atual, icnpj, cnpj_limpo, resultado, pasta_download, nome_empresa):

    logging.info(f"Processando CNPJ: {cnpj_limpo} - Parcelamento: {num_parcelamento}")

    if resultado is None:
        resultado = {
            "sucesso": False,
            "data_vencimento": None,
            "motivo_falha": None,
            "num_parcelamento": num_parcelamento
        }

    driver.get("https://www3.fazenda.sp.gov.br/CAWEB/Account/Login.aspx")
    time.sleep(5)

    contabilist = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/div[2]/div[1]/div/div/div/div[3]/div/div[1]/div/div[3]/div[1]/span/input[2]'))
    )
    contabilist.click()
    time.sleep(0.6)

    img_cert = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[3]/div[2]/div[1]/div/div/div/div[3]/div/div[2]/div/div[4]/input'))
    )
    img_cert.click()
    time.sleep(3)

    if selecionar_certificado(nome_arquivo_img=r"image_cert"):
        logging.info("Certificado selecionado com sucesso, continuando o processamento")

    else:
        logging.error("Falha so selecionar certificado, abortando operação")
        resultado["sucesso"] = False
        resultado["motivo_falha"] = "Falha ao selecionar certificado digital"
        return resultado
    
    time.sleep(3)

    cficmsp = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Conta Fiscal do ICMS e Parcelamento')]"))
    )
    cficmsp.click()
    time.sleep(1)

    parc = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Parcelamento')]"))
    )
    parc.click()
    time.sleep(1)

    coalt = WebDriverWait(driver, 5).until(
        EC.presence_of_element_located((By.XPATH, "//a[contains(text(), 'Consultar e Alterar')]"))
    )
    coalt.click()
    time.sleep(1)

    elemento_select = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[4]/div[2]/div[2]/div[5]/div/fieldset/div/select"))
    )
    select = Select(elemento_select)
    select.select_by_visible_text("CNPJ")

    campo_cnpj = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[4]/div[2]/div[2]/div[5]/div/fieldset/div/input"))
    )
    campo_cnpj.click()
    campo_cnpj.send_keys(Keys.HOME)
    time.sleep(0.5)

    logging.info(f"CNPJ encontrado: {cnpj_limpo}")

    for char in str(cnpj_limpo):
        campo_cnpj.send_keys(char)
        time.sleep(random.uniform(0.1, 0.2))

    time.sleep(1.5)

    consultar = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body/form/div[4]/div[2]/div[2]/div[5]/div/fieldset/div/a"))
    )
    consultar.click()
    time.sleep(1.5)

    linha_pedido = driver.find_elements(By.XPATH, "//table[@id='MainContent_gvListaPedidoParcelado']/tbody/tr")

    for l in range(len(linha_pedido)):
        linha_pedido_atual = l + 1

        situacao = driver.find_element(By.XPATH, f"//table[@id='MainContent_gvListaPedidoParcelado']/tbody/tr[{linha_pedido_atual}]/td[4]").text

        try:
                link = driver.find_element(By.XPATH, f"//table[@id='MainContent_gvListaPedidoParcelado']/tbody/tr[{linha_pedido_atual}]/td[1]/a")
                numero_pedido = link.text

                driver.execute_script("""
                    arguments[0].scrollIntoView({
                        block: 'center', 
                        behavior: 'smooth'
                    });
                """, link)
                
                time.sleep(2)

                logging.info(f"Clicando no parcelamento {numero_pedido} com situação '{situacao}'")
                link.click()

                try:
                    WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, "//table/tbody/tr"))
                    )

                except TimeoutException:
                    logging.error("A tabela de parcelas não carregou a tempo.")
                    resultado["sucesso"] = False
                    resultado["motivo_falha"] = "Tabela de parcelas não encontrada após clicar no parcelamento"
                    return resultado
                
                break
        
        except Exception as e:
            logging.warning(f"Erro ao processar linha {linha_pedido_atual}{str(e)[:100]}")
            continue
    
    resultado['atrasadas'] = "NÃO"
    parcelas_atrasadas = []

    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)

    baixou_guia = False

    try:
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)
        
        print("Voltou para o topo da página")

        tabela_parcelas = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/form/div[4]/div[2]/div[2]/div[5]/div/fieldset/div[8]/fieldset/div/table'))
        )
        linhas = tabela_parcelas.find_elements(By.TAG_NAME, 'tr')
        hoje = datetime.now()

        # NOVA LÓGICA: Buscar parcelas atrasadas, mês atual ou próxima disponível
        parcelas_atrasadas_download = []
        parcelas_disponiveis = []

        # Primeiro: catalogar TODAS as parcelas não liquidadas
        for linha in linhas:
            colunas = linha.find_elements(By.TAG_NAME, 'td')
            if len(colunas) < 3:
                continue

            numero_parcela = colunas[1].text.strip()
            liquidada = "**" in numero_parcela
            data_venc = colunas[2].text.strip()

            try:
                data_dt = datetime.strptime(data_venc, "%d/%m/%Y")
            except Exception:
                continue

            if not liquidada:
                parcela_info = {
                    'linha_elemento': linha,
                    'colunas': colunas,
                    'data': data_dt,
                    'data_texto': data_venc,
                    'mes_atual': data_dt.month == hoje.month and data_dt.year == hoje.year,
                    'atrasada': data_dt < hoje
                }
                
                if data_dt < hoje:
                    # Parcela atrasada
                    parcelas_atrasadas.append(data_venc)
                    parcelas_atrasadas_download.append(parcela_info)
                else:
                    # Parcela futura/atual
                    parcelas_disponiveis.append(parcela_info)

        # Segundo: escolher qual parcela baixar (PRIORIDADE: Atrasada > Mês atual > Próxima)
        parcela_escolhida = None

        if parcelas_atrasadas_download:
            # 1ª PRIORIDADE: Parcela mais antiga atrasada
            parcela_escolhida = min(parcelas_atrasadas_download, key=lambda x: x['data'])
            logging.info(f"PARCELA ATRASADA encontrada: {parcela_escolhida['data_texto']} - Baixando...")
            
        elif parcelas_disponiveis:
            # 2ª PRIORIDADE: Parcela do mês atual
            parcela_mes_atual = next((p for p in parcelas_disponiveis if p['mes_atual']), None)
            
            if parcela_mes_atual:
                parcela_escolhida = parcela_mes_atual
                logging.info(f"Parcela do mês atual encontrada: {parcela_escolhida['data_texto']}")
            else:
                # 3ª PRIORIDADE: Próxima parcela disponível
                parcela_escolhida = min(parcelas_disponiveis, key=lambda x: x['data'])
                logging.info(f"Próxima parcela disponível: {parcela_escolhida['data_texto']}")

        # Terceiro: baixar a parcela escolhida
        if parcela_escolhida and not baixou_guia:
            try:
                impressora = parcela_escolhida['colunas'][0].find_element(By.TAG_NAME, 'a')
                driver.execute_script("arguments[0].scrollIntoView({block: 'center', behavior: 'smooth'});", impressora)
                WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable(impressora)
                )
                driver.execute_script("arguments[0].click();", impressora)
                baixou_guia = True
                time.sleep(2)
                resultado["data_vencimento"] = parcela_escolhida['data_texto']

                if parcela_escolhida['atrasada']:
                    logging.info(f"Baixando DAE ATRASADO para {parcela_escolhida['data_texto']}")
                else:
                    logging.info(f"Baixando DAE para {parcela_escolhida['data_texto']}")
                
                arquivo_original = os.path.join(pasta_download, "DARE.pdf")
                max_espera = 5
                inicio = time.time()
                while time.time() - inicio < max_espera:
                    if os.path.exists(arquivo_original):
                        tamanho_inicial = os.path.getsize(arquivo_original)
                        time.sleep(1)
                        if os.path.getsize(arquivo_original) == tamanho_inicial:
                            break
                    time.sleep(1)

                if os.path.exists(arquivo_original):
                    pasta_base = PASTA_BASE
                    subpasta = "Parcelamentos ADM - SP"
                    pasta_destino = criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, data_venc)
                    nome_empresa_limpo = limpar_nome_arquivo(nome_empresa)
                    novo_nome = os.path.join(pasta_destino, f"DAE_{nome_empresa_limpo} - {num_parcelamento}.pdf")
                    
                    if os.path.exists(novo_nome):
                        os.remove(novo_nome)
                    shutil.move(arquivo_original, novo_nome)
                    logging.info(f"Arquivo renomeado e movido para: {novo_nome}")
                else:
                    logging.warning(f"Arquivo de download não encontrado após {max_espera} segundos")
                    resultado["motivo_falha"] = "Arquivo de download não encontrado"

                resultado["data_vencimento"] = parcela_escolhida['data_texto']
                resultado["sucesso"] = True
                guia_encontrada = True

            except Exception as e:
                resultado["motivo_falha"] = f'Erro ao clicar na impressora: {str(e)[:60]}'
                logging.error(resultado["motivo_falha"])

    except Exception as e:
        resultado["motivo_falha"] = f'Erro ao clicar na impressora: {str(e)[:60]}'
        logging.error(resultado["motivo_falha"])

    if parcelas_atrasadas:
        resultado['atrasadas'] = f"SIM - {', '.join(parcelas_atrasadas)}"

    else:
        resultado['atrasadas'] = "NÃO"

    if not baixou_guia:
        resultado["motivo_falha"] = "Não encontrou parcela do mês atual para baixar"
    
    if resultado is None:

        resultado = {
            "num_parcelamento": num_parcelamento,
            "nome_empresa": nome_empresa,
            "sucesso": False,
            "data_vencimento": None,
            "motivo_falha": "Erro desconhecido"
        }

    return resultado

def obter_penultimo_dia_util():

    data_atual = datetime.now()
    ano_atual = data_atual.year
    mes_atual = data_atual.month

    _, ultimo_dia = calendar.monthrange(ano_atual, mes_atual)
    todos_dias = [datetime(ano_atual, mes_atual, dia) for dia in range(1, ultimo_dia + 1)]

    feriados_br = holidays.Brazil(years=ano_atual)

    dias_uteis = [dia for dia in todos_dias if dia.weekday() < 5 and dia.strftime('%Y-%m-%d') not in feriados_br]

    penultimo_dia_util = dias_uteis[-2]

    return penultimo_dia_util.strftime('%d%m%Y')

def pintar_planilha(caminho_planilha, resultado):
    try:
        logging.info(f"Iniciando processo de coloração da planilha: {caminho_planilha}")

        wb = openpyxl.load_workbook(caminho_planilha)

        data_hoje = datetime.now()
        mes_atual = data_hoje.month
        ano_atual = data_hoje.year

        meses = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
        }
        nome_aba = f"{meses[mes_atual]} {ano_atual}"

        if nome_aba not in wb.sheetnames:
            logging.error(f"Aba {nome_aba} não encontrada na planilha")
            return False
        
        ws = wb[nome_aba]

        coluna_situacao_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Situação":
                coluna_situacao_idx = col_idx
                break
        
        if coluna_situacao_idx is None:
            coluna_situacao_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_situacao_idx).value = "Situação"
            ws.cell(row=1, column=coluna_situacao_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Situação' criada na posição {coluna_situacao_idx}")

        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        azul = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        mes_atual_num = f"{mes_atual:02d}/{ano_atual}"

        num_parcelamento = resultado["num_parcelamento"]
        sucesso = resultado["sucesso"]
        data_vencimento = resultado["data_vencimento"]
        motivo_falha = resultado.get("motivo_falha", "")
        situacao_especial = resultado.get("situacao_especial", "")

        linha_encontrada = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[5].value).strip() == str(num_parcelamento).strip():
                linha_encontrada = row_idx
                break
            
        if not linha_encontrada:
            logging.warning(f"Linha para parcelamento {num_parcelamento} não encontrada na planilha")
            return False

        cor_a_aplicar = None
        texto_situacao = ""

        if sucesso:
            try:
                partes_data = data_vencimento.split('/')

                meses_str_para_num = {
                    'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
                    'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
                    'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'
                }
                mes_raw = partes_data[1].strip().upper()

                if mes_raw.isdigit():
                    mes_site = f"{int(mes_raw):02d}/{partes_data[2]}"
                
                else:
                    mes_abrev = mes_raw[:3]
                    mes_site = f"{meses_str_para_num.get(mes_abrev, '00')}/{partes_data[2]}"

                if mes_site == mes_atual_num:
                    cor_a_aplicar = verde
                    texto_situacao = "OK"
                    logging.info(f"Pintado linha {linha_encontrada} de VERDE (data correspondente)")
                    
                else:
                    cor_a_aplicar = amarelo
                    texto_situacao = f"OK / Data Divergente ({data_vencimento})"
                    logging.info(f"Pintando linha {linha_encontrada} de AMARELO (data diferente: {mes_site} vs {mes_atual_num})")
                
            except:
                cor_a_aplicar = amarelo
                texto_situacao = f"OK / Data Divergente (formato inválido: {data_vencimento})"
                logging.warning(f"Formato de data inesperado: {data_vencimento}, pintado de AMARELO")
            
        elif situacao_especial == "quitado":
            cor_a_aplicar = azul
            texto_situacao = "Parcelamento Quitado"
            logging.info(f"Pintando linha {linha_encontrada} de AZUL (parcelamento quitado)")
            
        elif situacao_especial == "desistente":
            cor_a_aplicar = vermelho
            texto_situacao = "Parcelamento Desistente"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (parcelamento desistente)")
            
        else:
            cor_a_aplicar = vermelho
            texto_situacao = f"ERRO: {motivo_falha}"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (falha no processamento)")

        cell = ws.cell(row=linha_encontrada, column=coluna_situacao_idx)
        cell.value = texto_situacao
        
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if linha_encontrada == 2:
            col_letter = get_column_letter(coluna_situacao_idx)
            ws.column_dimensions[col_letter].width = 40
        
        coluna_atrasadas_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Atrasadas?":
                coluna_atrasadas_idx = col_idx
                break
        
        if coluna_atrasadas_idx is None:
            coluna_atrasadas_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_atrasadas_idx).value = "Atrasadas?"
            ws.cell(row=1, column=coluna_atrasadas_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Atrasadas?' criada na posição {coluna_atrasadas_idx}")
        
        valor_atrasadas = resultado.get("atrasadas", "NÃO")
        ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).value = valor_atrasadas

        total_colunas = ws.max_column
        for col_idx in range(1, total_colunas + 1):
            ws.cell(row=linha_encontrada, column=col_idx).fill = cor_a_aplicar
        
        cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if valor_atrasadas.startswith("SIM"):
            ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).fill=cinza

        ajustar_largura_colunas(ws)

        wb.save(caminho_planilha)
        logging.info(f"Linha {linha_encontrada} colorida, situação atualizada e planilha salva com sucesso")
        wb.close()
        return True
    
    except Exception as e:
        logging.error(f"Erro ao pintar linha planilha: {e}")
        logging.error(traceback.format_exc())
        return False
    
def main():
    try:
        caminho_certificado = r"Path_Certificado"
        nome_certificado =  "Nome_Certificado"

        driver, pasta_download = configurar_navegador_com_certificado(caminho_certificado, nome_certificado)
        
        dados, coluna_tipo, coluna_numero, mes_atual_texto = carregar_dados_planilha_mensal()
        
        if not dados is None and not coluna_tipo is None and not coluna_numero is None:
            logging.info(f"Dados carregados com sucesso da planilha-mãe")           
            
            dados_filtrados = dados.dropna(subset=[coluna_tipo])
            dados_filtrados = dados_filtrados[dados_filtrados[coluna_tipo].str.lower().str.strip().str.contains(r'estadual\s*-?\s*adm\s*-?\s*sp\b', regex=True, na=False)]
            logging.info(f"Total de registros filtrados: {len(dados_filtrados)}")

            if len(dados_filtrados) == 0:
                logging.warning("Nenhum número de parcelamento 'ESTADUAL ADM - SP' encontrado na planilha!")
                return
            
            numeros_parcelamento = dados_filtrados[coluna_numero].dropna().unique().tolist()
            logging.info(f"Quantidade de parcelamentos encontrados: {len(numeros_parcelamento)}")
            logging.info(f"Números de parcelamentos encontrados: {numeros_parcelamento}")

            caminho_planilha = CAMINHO_PLANILHA

            for _, row in dados_filtrados.iterrows():
                try:
                    num_parcelamento = row[coluna_numero]
                    nome_empresa = row['EMPRESA']
                    coluna_cnpj = None

                    possiveis_nomes = ['CNPJ/CPF/CAEPF/CNO', 'CNPJ FORMATADO', 'CNPJ EMPRESA']

                    for nome in possiveis_nomes:
                        if nome in row:
                            coluna_cnpj = nome
                            break

                    if not coluna_cnpj:
                        logging.warning("Nenhuma coluna de CNPJ encontrada na linha. Verifique os nomes das colunas.")
                        cnpj_raw = ''

                    else:
                        cnpj_raw = row[coluna_cnpj]

                    logging.info(f"CNPJ bruto da planilha: {cnpj_raw}")
                    cnpj_empresa = limpar_cnpj(cnpj_raw)
                    logging.info(f"CNPJ limpo: {cnpj_empresa}")

                    logging.info(f"Processando parcelamento: {num_parcelamento} - Empresa: {nome_empresa}")

                    resultado = {
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": None,
                        "num_parcelamento": num_parcelamento,
                        "nome_empresa": nome_empresa
                    }

                    resultado = baixar_dae_mes_atual(driver, num_parcelamento, mes_atual_texto, datetime.now().year, 1, cnpj_empresa, resultado, pasta_download, nome_empresa)

                    resultado["num_parcelamento"] = num_parcelamento
                    resultado["nome_empresa"] = nome_empresa

                    if resultado["sucesso"]:
                        logging.info(f"Processamento bem-sucedido para parcelamento {num_parcelamento}")
                    
                    else:
                        logging.warning(f"Falha no processamento do parcelamento {num_parcelamento}: {resultado['motivo_falha']}")

                    if pintar_planilha(caminho_planilha, resultado):
                        logging.info(f"Linha para parcelamento {num_parcelamento} colorida com sucesso")

                    else:
                        logging.error(f"Falha ao colorir linha para parcelamento {num_parcelamento}")                     

                    time.sleep(random.uniform(3,5))
                
                except Exception as e:
                    logging.error(f"Erro ao processar linha para parcelamento: {e}")
                    resultado = {
                        "num_parcelamento": num_parcelamento if 'num_parcelamento' in locals() else "Desconhecido",
                        "nome_empresa": nome_empresa if 'nome_empresa' in locals() else "Desconhecido",
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": str(e)[:100]
                    }
                    pintar_planilha(caminho_planilha, resultado)
                    continue
            
            logging.info(f"Processamento concluído para {len(dados_filtrados)} parcelamentos")
        
        else:
            logging.error("Não foi possível carregar os dados da planilha-mãe")
    
    except Exception as e:
        logging.error(f"Erro no processo principal: {e}")
    
    finally:
        try:
            driver.quit()
            logging.info("Navegador fechado")
            
        except:
            pass

if __name__ == "__main__":
    try:
        logging.info("Iniciando script de automação de parcelamentos")
        main()
        logging.info("Script finalizado com sucesso")
    
    except Exception as e:
        logging.error(f"Erro fatal: {e}")
    
    print("\nProcessamento concluído!")
    sys.exit(0)
