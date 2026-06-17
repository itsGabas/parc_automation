import undetected_chromedriver as uc
import re
import pandas as pd
import sys
import time
import os
import random
import openpyxl
import logging
import shutil
import traceback
import math
import selenium_stealth
import unicodedata
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from pathlib import Path
from utils import (
    carregar_dados_planilha_mensal,
    ajustar_largura_colunas,
    click_element_safely,
    extrair_numero_cda,
    extrair_numero_parcelamento,
    CAMINHO_PLANILHA,
    PASTA_DOWNLOADS,
    PASTA_BASE,
    MESES_ABREV,
    MESES_NUMERO,
)

def normaliza(txt):
    return unicodedata.normalize('NFKD', txt.lower()).encode('ascii', 'ignore').decode('ascii')

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)

def configurar_navegador():
    pasta_download_base = PASTA_BASE
    subpasta = "Parcelamentos D.A. - SP"
    pasta_download = criar_estrutura_pastas_por_vencimento(pasta_download_base, subpasta)

    logging.info(f"Usando pasta de download: {pasta_download}")

    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-popup-blocking")

    prefs = {
        "download.default_directory": pasta_download,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_setting_values.popups": 1

    }
    options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    selenium_stealth.stealth(
        driver,
        languages=["pt-BR", "pt"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,
    )

    params = {
        "behavior": "allow",
        "downloadPath": pasta_download
    }
    driver.execute_cdp_cmd("Page.setDownloadBehavior", params)

    return driver, pasta_download

def limpar_nome_arquivo(nome):
    caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']

    nome_limpo = nome
    for char in caracteres_invalidos:
        nome_limpo = nome_limpo.replace(char, '_')
    
    if len(nome_limpo) > 50:
        nome_limpo = nome_limpo[:47] + '...'
    
    return nome_limpo.strip()

def criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, data_vencimento=None):
    meses_abrev = {
        1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR',
        5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
        9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
    }
    
    if data_vencimento:
        try:
            partes_data = data_vencimento.split('/')
            
            mes_raw = partes_data[1].strip().upper()
            ano = partes_data[2].strip()
            
            if mes_raw.isdigit():
                mes_num = int(mes_raw)

            else:
                meses_str_para_num = {
                    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4,
                    'MAI': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
                    'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
                }
                mes_abrev = mes_raw[:3]
                mes_num = meses_str_para_num.get(mes_abrev, datetime.now().month)
            
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
            
        except Exception as e:
            logging.warning(f"Erro ao processar data de vencimento '{data_vencimento}', usando mês atual: {str(e)}")
            data_atual = datetime.now()
            mes_num = data_atual.month
            ano = str(data_atual.year)
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"

    else:
        data_atual = datetime.now()
        mes_num = data_atual.month
        ano = str(data_atual.year)
        mes_abrev_pasta = meses_abrev[mes_num]
        meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
        numero_mes = meses_dict[mes_abrev_pasta.upper()]
        nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
    
    caminho_pasta_mes = os.path.join(pasta_base, nome_pasta_mes)
    
    if not os.path.exists(caminho_pasta_mes):
        os.makedirs(caminho_pasta_mes)
        logging.info(f"Pasta do mês criada: {caminho_pasta_mes}")
    
    caminho_completo = os.path.join(caminho_pasta_mes, subpasta)
    
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
        logging.info(f"Subpasta criada: {caminho_completo}")
    
    return caminho_completo

def baixar_dae_mes_atual(driver, num_cda, mes_atual_texto, ano_atual, icnpj, pasta_download, resultado, num_parcelamento, nome_empresa, numero_mes, mes_abrev_pasta):
    try:
        logging.info(f"Iniciando processo para CDA {num_cda} e CNPJ {icnpj}")

        driver.get("https://www.dividaativa.pge.sp.gov.br/sc/pages/pagamento/gareParcelamento.jsf")
        time.sleep(random.uniform(2, 5))

        try:
            cda = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/table/tbody/tr/td/div/div[2]/table[2]/tbody/tr[3]/td/input')
            cda.clear()
            
            for char in str(num_cda):
                cda.send_keys(char)
                time.sleep(random.uniform(0.1, 0.2))
            
            time.sleep(1)

            avancar = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/div[1]/input')
            click_element_safely(driver, avancar)
            time.sleep(1)
            
            try:
                parcelamento_site_element = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/table/tbody/tr/td/div/div[2]/table[2]/tbody/tr/td[2]')
                parcelamento_site = parcelamento_site_element.text.strip()
                
                logging.info(f"Número de parcelamento encontrado no site: {parcelamento_site}")

                try:
                    avancar2 = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/div[1]/input[2]')
                    click_element_safely(driver, avancar2)
                    time.sleep(1)

                    avancar3 = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/div[1]/input[2]')
                    click_element_safely(driver, avancar3)
                    time.sleep(1)

                    try:
                        tabela_parcelas = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/table/tbody/tr/td/div/div[2]/table[2]')
                        linhas = tabela_parcelas.find_elements(By.TAG_NAME, 'tr')
                        parcelas_atrasadas = []
                        data_vencimento_atual = None

                        for linha in linhas:
                            colunas = linha.find_elements(By.TAG_NAME, "td")
                            if len(colunas) < 4:
                                continue

                            data_venc = colunas[2].text.strip()
                            situacao = normaliza(colunas[3].text.strip())
                            situacao_original = colunas[3].text.strip()
                            logging.info(f"Lendo linha: vencimento='{data_venc}', situacao='{situacao_original}' (normalizado='{situacao}')")

                            if "atrasado" in situacao or "nao pago" in situacao:
                                parcelas_atrasadas.append(data_venc)

                            if "aguardando pagamento" in situacao:
                                data_vencimento_atual = data_venc

                        if parcelas_atrasadas:
                            resultado["atrasadas"] = f"SIM - {', '.join(parcelas_atrasadas)}"
                            logging.info("Parcelas atrasadas encontradas!")                         

                        else:
                            resultado["atrasadas"] = "NÃO"
                            logging.info("Nenhuma parcela atrasada encontrada!")    

                        if data_vencimento_atual:
                            resultado["data_vencimento"] = data_vencimento_atual
                            logging.info(f"Data de vencimento encontrada: {data_vencimento_atual}")

                        else:
                            data_hoje = datetime.now()
                            resultado["data_vencimento"] = f"15/{data_hoje.month:02d}/{data_hoje.year}"
                            logging.warning(f"Data de vencimento não encontrada, usando padrão: {resultado['data_vencimento']}")

                    except Exception as e:
                        resultado["atrasadas"] = "NÃO"
                        data_hoje = datetime.now()
                        resultado["data_vencimento"] = f"15/{data_hoje.month:02d}/{data_hoje.year}"
                        logging.warning(f"Erro ao processar tabela de parcelas: {e}")
                    
                    logging.debug(resultado["atrasadas"])

                    avancar4 = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/div[1]/input[2]')
                    click_element_safely(driver, avancar4)
                    time.sleep(2)

                    try:     
                        xis = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/div/img')
                        xis.click()
                        logging.info("Elemento X encontrado e clicado.")
                        time.sleep(1)
                    
                    except (TimeoutException, NoSuchElementException):
                        logging.info("Elemanto X não encontrado, continunando execução...")

                    download = driver.find_element(By.XPATH, '/html/body/div/div/div/div/div[2]/div/div[3]/div/div[2]/div[2]/span/form/table/tbody/tr/td/div/div[2]/div[1]/input')
                    click_element_safely(driver, download)
                    time.sleep(11)

                    download2 = driver.find_element(By.XPATH, '/html/body/div[2]/div[2]/div/div[2]/table/tbody/tr[2]/td/div[2]/input')
                    click_element_safely(driver, download2)
                    logging.info("Clicado no botão de download da GARE")
                    time.sleep(2)

                    arquivo_original = os.path.join(pasta_download, "DARE.pdf")
                    max_espera = 5
                    inicio = time.time()

                    while time.time() - inicio < max_espera:
                        if os.path.exists(arquivo_original):
                            tamanho_inicial = os.path.getsize(arquivo_original)
                            time.sleep(1)
                            if os.path.getsize(arquivo_original) == tamanho_inicial:
                                break
                        time.sleep(1)
                    
                    if os.path.exists(arquivo_original):
                        pasta_base = PASTA_BASE
                        data_vencimento = resultado["data_vencimento"]

                        partes_data = data_vencimento.split('/')
                        meses_abrev = {
                            '01': 'JAN', '02': 'FEV', '03': 'MAR', '04': 'ABR',
                            '05': 'MAI', '06': 'JUN', '07': 'JUL', '08': 'AGO',
                            '09': 'SET', '10': 'OUT', '11': 'NOV', '12': 'DEZ'
                        }
                        mes = meses_abrev[partes_data[1].zfill(2)]
                        ano = partes_data[2]
                        nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"

                        subpasta = "Parcelamentos D.A. - SP"
                        pasta_destino = os.path.join(pasta_base, nome_pasta_mes, subpasta)
                        os.makedirs(pasta_destino, exist_ok=True)

                        nome_empresa_limpo = limpar_nome_arquivo(resultado["nome_empresa"])
                        novo_nome = os.path.join(pasta_destino, f"DARE_{nome_empresa_limpo} - {num_cda}.pdf")

                        if os.path.exists(novo_nome):
                            os.remove(novo_nome)
                        
                        shutil.move(arquivo_original, novo_nome)
                        logging.info(f"Arquivo renomeado e movido para: {novo_nome}")
                        resultado["sucesso"] = True
                    
                    else:
                        logging.warning(f"Arquivo de download não encontrado após {max_espera} segundos")
                        resultado["motivo_falha"] = "Arquivo de download não encontrado"

                    resultado["data_vencimento"] = data_vencimento
                    resultado["sucesso"] = True
                    guia_encontrada = True
                    
                except Exception as e:
                    resultado["atrasadas"] = "NÃO"
                    logging.error(f"Erro ao tentar emitir GARE: {e}")
                    
            except NoSuchElementException:
                logging.warning(f"Não foi possível encontrar informações do parcelamento para o CDA {num_cda}")
                return {
                    "sucesso": False,
                    "data_vencimento": resultado.get("data_vencimento", None),
                    "motivo_falha": "Informações do parcelamento não encontradas",
                    "num_cda": num_cda,
                }
        
        except Exception as e:
            logging.error(f"Erro ao tentar preencher o campo com o número do CDA: {e}")
            return {
                "sucesso": False,
                "data_vencimento": resultado.get("data_vencimento", None),
                "motivo_falha": str(e)[:100],
                "num_cda": num_cda,
            }
    
    except Exception as e:
        logging.error(f"Erro ao tentar entrar no site: {e}")
        return {"sucesso": False, "data_vencimento": None, "motivo_falha": str(e)[:100]}
    
    if resultado is None:
        resultado = {
            "num_parcelamento": num_parcelamento,
            "nome_empresa": nome_empresa,
            "sucesso": False,
            "data_vencimento": None,
            "motivo_falha": "Erro desconhecido"
        }
    return resultado

def pintar_planilha(caminho_planilha, resultado, nome_empresa, num_cda):
    try:
        logging.info(f"Iniciando processo de coloração da planilha: {caminho_planilha}")

        wb = openpyxl.load_workbook(caminho_planilha)

        data_hoje = datetime.now()
        mes_atual = data_hoje.month
        ano_atual = data_hoje.year

        meses = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
        }
        nome_aba = f"{meses[mes_atual]} {ano_atual}"

        if nome_aba not in wb.sheetnames:
            logging.error(f"Aba {nome_aba} não encontrada na planilha")
            return {
                "sucesso": False,
                "data_vencimento": resultado.get("data_vencimento", None),
                "motivo_falha": "Aba não encontrada",
                "num_parcelamento": resultado.get("num_parcelamento", "Desconhecido"),
                "nome_empresa": nome_empresa
            }
        
        ws = wb[nome_aba]

        coluna_situacao_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Situação":
                coluna_situacao_idx = col_idx
                break
        
        if coluna_situacao_idx is None:
            coluna_situacao_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_situacao_idx).value = "Situação"
            ws.cell(row=1, column=coluna_situacao_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Situação' criada na posição {coluna_situacao_idx}")

        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        azul = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        mes_atual_num = f"{mes_atual:02d}/{ano_atual}"

        num_parcelamento = resultado.get("num_parcelamento", "Desconhecido")
        sucesso = resultado.get("sucesso", False)
        data_vencimento = resultado.get("data_vencimento", None)
        motivo_falha = resultado.get("motivo_falha", "")
        situacao_especial = resultado.get("situacao_especial", "")

        linha_encontrada = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            # Buscar pela CDA (que é o que realmente identifica a linha)
            valor_celula = str(row[5].value or "").strip()
            
            # Tentar buscar pela CDA primeiro
            if num_cda and str(num_cda).strip() in valor_celula:
                linha_encontrada = row_idx
                logging.info(f"Linha encontrada pela CDA {num_cda} na linha {row_idx}")
                break
            
            # Se não encontrar pela CDA, tentar pelo parcelamento (caso original)
            elif num_parcelamento and str(num_parcelamento).strip() != "Desconhecido" and str(num_parcelamento).strip() in valor_celula:
                linha_encontrada = row_idx
                logging.info(f"Linha encontrada pelo parcelamento {num_parcelamento} na linha {row_idx}")
                break
            
        if not linha_encontrada:
            logging.warning(f"Linha para parcelamento {num_parcelamento} não encontrada na planilha")
            return {
                "sucesso": False,
                "data_vencimento": resultado.get("data_vencimento", None),
                "motivo_falha": "Aba não encontrada",
                "num_parcelamento": resultado.get("num_parcelamento", "Desconhecido"),
                "nome_empresa": nome_empresa
            }

        cor_a_aplicar = None
        texto_situacao = ""

        if sucesso:
            try:
                partes_data = data_vencimento.split('/')

                meses_str_para_num = {
                    'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
                    'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
                    'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'
                }
                mes_raw = partes_data[1].strip().upper()

                if mes_raw.isdigit():
                    mes_site = f"{int(mes_raw):02d}/{partes_data[2]}"
                
                else:
                    mes_abrev = mes_raw[:3]
                    mes_site = f"{meses_str_para_num.get(mes_abrev, '00')}/{partes_data[2]}"

                if mes_site == mes_atual_num:
                    cor_a_aplicar = verde
                    texto_situacao = "OK"
                    logging.info(f"Pintado linha {linha_encontrada} de VERDE (data correspondente)")
                    
                else:
                    cor_a_aplicar = amarelo
                    texto_situacao = f"OK / Data Divergente ({data_vencimento})"
                    logging.info(f"Pintando linha {linha_encontrada} de AMARELO (data diferente: {mes_site} vs {mes_atual_num})")
                
            except:
                cor_a_aplicar = amarelo
                texto_situacao = f"OK / Data Divergente (formato inválido: {data_vencimento})"
                logging.warning(f"Formato de data inesperado: {data_vencimento}, pintado de AMARELO")
            
        elif situacao_especial == "quitado":
            cor_a_aplicar = azul
            texto_situacao = "Parcelamento Quitado"
            logging.info(f"Pintando linha {linha_encontrada} de AZUL (parcelamento quitado)")
            
        elif situacao_especial == "desistente":
            cor_a_aplicar = vermelho
            texto_situacao = "Parcelamento Desistente"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (parcelamento desistente)")
            
        else:
            cor_a_aplicar = vermelho
            texto_situacao = f"ERRO: {motivo_falha}"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (falha no processamento)")

        cell = ws.cell(row=linha_encontrada, column=coluna_situacao_idx)
        cell.value = texto_situacao
        
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if linha_encontrada == 2:
            col_letter = get_column_letter(coluna_situacao_idx)
            ws.column_dimensions[col_letter].width = 40
        
        coluna_atrasadas_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Atrasadas?":
                coluna_atrasadas_idx = col_idx
                break
        
        if coluna_atrasadas_idx is None:
            coluna_atrasadas_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_atrasadas_idx).value = "Atrasadas?"
            ws.cell(row=1, column=coluna_atrasadas_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Atrasadas?' criada na posição {coluna_atrasadas_idx}")
        
        valor_atrasadas = resultado.get("atrasadas", "NÃO")
        ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).value = valor_atrasadas

        total_colunas = ws.max_column
        for col_idx in range(1, total_colunas + 1):
            ws.cell(row=linha_encontrada, column=col_idx).fill = cor_a_aplicar
        
        cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if valor_atrasadas.startswith("SIM"):
            ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).fill=cinza
        
        ajustar_largura_colunas(ws)

        wb.save(caminho_planilha)
        logging.info(f"Linha {linha_encontrada} colorida, situação atualizada e planilha salva com sucesso")
        return {
            "sucesso": True,
            "data_vencimento": resultado.get("data_vencimento"),
            "motivo_falha": "",
            "num_parcelamento": resultado.get("num_parcelamento"),
            "nome_empresa": nome_empresa
        }
    
    except Exception as e:
        logging.error(f"Erro ao pintar linha planilha: {e}")
        logging.error(traceback.format_exc())
        return {
            "sucesso": False,
            "data_vencimento": resultado.get("data_vencimento"),
            "motivo_falha": "str(e)[:100]",
            "num_parcelamento": resultado.get("num_parcelamento"),
            "nome_empresa": nome_empresa
        }
    
def main():
    try:
        driver, pasta_download = configurar_navegador()
        
        dados, coluna_tipo, coluna_numero, mes_atual_texto = carregar_dados_planilha_mensal()
        
        if not dados is None and not coluna_tipo is None and not coluna_numero is None:
            logging.info(f"Dados carregados com sucesso da planilha-mãe")           
            
            dados_filtrados = dados.dropna(subset=[coluna_tipo])
            dados_filtrados = dados_filtrados[dados_filtrados[coluna_tipo].str.lower().str.strip().str.contains(r'estadual\s*d\.?\s*a\.?\s*-?\s*sp', regex=True)]
            logging.info(f"Total de registros filtrados: {len(dados_filtrados)}")

            if len(dados_filtrados) == 0:
                logging.warning("Nenhum número de parcelamento 'ESTADUAL D.A. - SP' encontrado na planilha!")
                return
            
            dados_filtrados['Num_Parcelamento_Limpo'] = dados_filtrados[coluna_numero].apply(extrair_numero_parcelamento)
            dados_filtrados['CDA'] = dados_filtrados[coluna_numero].apply(extrair_numero_cda)

            data_atual = datetime.now()
            mes_num = data_atual.month
            ano = str(data_atual.year)
            
            meses_abrev = {
                1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR',
                5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
                9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
            }
            mes_abrev_pasta = meses_abrev[mes_num]
            
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]

            caminho_planilha = CAMINHO_PLANILHA

            for _, row in dados_filtrados.iterrows():
                try:
                    num_parcelamento_completo = row[coluna_numero]
                    num_parcelamento = row['Num_Parcelamento_Limpo']
                    num_cda = row['CDA']
                    nome_empresa = row['EMPRESA']

                    if pd.isna(num_cda):
                        logging.warning(f"CDA não encontrado para o parcelamento {num_parcelamento}. Pulando...")
                        continue

                    logging.info(f"Processando CDA: {num_cda} - Parcelamento: {num_parcelamento} - Empresa: {nome_empresa}")

                    icnpj = row.get('CNPJ/CPF/CAEPF/CNO', '')

                    resultado = {
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": None,
                        "num_parcelamento": num_parcelamento,
                        "nome_empresa": nome_empresa
                    }

                    resultado = baixar_dae_mes_atual(driver, num_cda, mes_atual_texto, datetime.now().year, icnpj, pasta_download, resultado, num_parcelamento, nome_empresa, numero_mes, mes_abrev_pasta)

                    if not resultado or not isinstance(resultado, dict):
                        resultado = {
                            "num_parcelamento": num_parcelamento,
                            "nome_empresa": nome_empresa,
                            "sucesso": False,
                            "data_vencimento": None,
                            "motivo_falha": "Função não retornou resultado válido"
                        }

                    if resultado.get("sucesso"):
                        logging.info(f"Processamento bem-sucedido para parcelamento {num_parcelamento}")

                    else:
                        logging.warning(f"Falha no processamento do parcelamento {num_parcelamento}: {resultado.get('motivo_falha')}")

                    resultado_pintura = pintar_planilha(caminho_planilha, resultado, nome_empresa, num_cda)

                    time.sleep(random.uniform(3,5))
                
                except Exception as e:
                    logging.error(f"Erro ao processar linha para parcelamento: {e}")
                    resultado = {
                        "num_parcelamento": locals().get('num_parcelamento', 'Desconhecido'),
                        "nome_empresa": locals().get('nome_empresa', 'Desconhecido'),
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": str(e)[:100]
                    }

                    if resultado["nome_empresa"] and resultado["num_parcelamento"]:
                        pintar_planilha(caminho_planilha, resultado, resultado["nome_empresa"], num_cda)
                        
                    else:
                        logging.error("Dados insuficientes para pintar a planilha.")

                    continue
            
            logging.info(f"Processamento concluído para {len(dados_filtrados)} parcelamentos")
        
        else:
            logging.error("Não foi possível carregar os dados da planilha-mãe")
    
    except Exception as e:
        logging.error(f"Erro no processo principal: {e}")
    
    finally:
        try:
            if 'driver' in locals() and driver is not None:
                try:
                    driver.quit()
                except Exception as e:
                    logging.warning(f"Erro ao encerrar WebDriver: {e}")
                finally:
                    driver = None
            logging.info("Navegador fechado")
        except Exception:
            pass

if __name__ == "__main__":
    try:
        logging.info("Iniciando script de automação de parcelamentos")
        main()
        logging.info("Script finalizado com sucesso")
    
    except Exception as e:
        logging.error(f"Erro fatal: {e}")
    
    print("\nProcessamento concluído!")
    sys.exit(0)
