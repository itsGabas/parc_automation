import undetected_chromedriver as uc
import re
import pandas as pd
import sys
import time
import os
import random
import openpyxl
import logging
import shutil
import traceback
import math
import calendar
import holidays
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from pathlib import Path
from utils import (
    carregar_dados_planilha_mensal,
    ajustar_largura_colunas,
    extrair_numero_cda,
    extrair_numero_parcelamento,
    CAMINHO_PLANILHA,
    PASTA_BASE,
    PASTA_DOWNLOADS,
    MESES_ABREV,
    MESES_NUMERO,
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)

def penultimo_dia_util():
    data_atual = datetime.now()
    ano_atual = data_atual.year
    mes_atual = data_atual.month

    _, ultimo_dia = calendar.monthrange(ano_atual, mes_atual)
    todos_dias = [datetime(ano_atual, mes_atual, dia) for dia in range(1, ultimo_dia +1)]

    feriados_br = holidays.Brazil(years=ano_atual)

    dias_uteis = [dia for dia in todos_dias if dia.weekday() < 5 and dia not in feriados_br]

    penultimo_dia_util = dias_uteis[-2]

    return penultimo_dia_util.strftime('%d%m%Y')

def extrair_numero_cda(texto):
    try:
        if pd.isna(texto):
            return None
        
        texto = str(texto)
        
        match = re.search(r'CDA:?\s*(\d+)', texto, re.IGNORECASE)

        if match:
            return match.group(1)
        
        if texto.strip().isdigit():
            return texto.strip()
        
        return texto

    except Exception as e:
        logging.error(f"Erro ao extrair número CDA: {e}")
        return texto
    
def extrair_numero_parcelamento(texto):
    try:
        if pd.isna(texto):
            return None

        texto = str(texto)

        match = re.search(r'^(\d+)\s*/\s*CDA', texto)
        if match:
            return match.group(1)
        
        return texto
    
    except Exception as e:
        logging.error(f"Erro ao extrair número do parcelamento: {e}")
        return texto

def configurar_navegador():
    pasta_download_base = PASTA_BASE
    subpasta = "Parcelamentos - RS"
    pasta_download = criar_estrutura_pastas_por_vencimento(pasta_download_base, subpasta)

    logging.info(f"Usando pasta de download: {pasta_download}")

    options = uc.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["disable-popup-blocking"])
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-infobars")

    prefs = {
        "download.default_directory": pasta_download,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_setting_values.popups": 1

    }
    options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)

    params = {
        "behavior": "allow",
        "downloadPath": pasta_download
    }
    driver.execute_cdp_cmd("Page.setDownloadBehavior", params)

    return driver, pasta_download

def limpar_nome_arquivo(nome):
    caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']

    nome_limpo = nome
    for char in caracteres_invalidos:
        nome_limpo = nome_limpo.replace(char, '_')
    
    if len(nome_limpo) > 50:
        nome_limpo = nome_limpo[:47] + '...'
    
    return nome_limpo.strip()

def limpar_cnpj(cnpj_sujo):
    if pd.isna(cnpj_sujo) or cnpj_sujo == '':
        return ''
    
    try:
        cnpj_str = str(int(float(cnpj_sujo)))
    
    except:
        cnpj_str = str(cnpj_sujo)
    
    apenas_numeros = ''.join(char for char in cnpj_str if char.isdigit())

    return apenas_numeros.zfill(14)

def criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, data_vencimento=None):
    meses_abrev = {
        1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR',
        5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
        9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
    }
    
    if data_vencimento:
        try:
            partes_data = data_vencimento.split('/')
            
            mes_raw = partes_data[1].strip().upper()
            ano = partes_data[2].strip()
            
            if mes_raw.isdigit():
                mes_num = int(mes_raw)

            else:
                meses_str_para_num = {
                    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4,
                    'MAI': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
                    'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
                }
                mes_abrev = mes_raw[:3]
                mes_num = meses_str_para_num.get(mes_abrev, datetime.now().month)
            
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
            
        except Exception as e:
            logging.warning(f"Erro ao processar data de vencimento '{data_vencimento}', usando mês atual: {str(e)}")
            data_atual = datetime.now()
            mes_num = data_atual.month
            ano = str(data_atual.year)
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"

    else:
        data_atual = datetime.now()
        mes_num = data_atual.month
        ano = str(data_atual.year)
        mes_abrev_pasta = meses_abrev[mes_num]
        meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
        numero_mes = meses_dict[mes_abrev_pasta.upper()]
        nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
    
    caminho_pasta_mes = os.path.join(pasta_base, nome_pasta_mes)
    
    if not os.path.exists(caminho_pasta_mes):
        os.makedirs(caminho_pasta_mes)
        logging.info(f"Pasta do mês criada: {caminho_pasta_mes}")
    
    caminho_completo = os.path.join(caminho_pasta_mes, subpasta)
    
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
        logging.info(f"Subpasta criada: {caminho_completo}")
    
    return caminho_completo

def baixar_dae_mes_atual(driver, num_cda, mes_atual_texto, ano_atual, icnpj, pasta_download, resultado):
    try:
        logging.info(f"Iniciando processo para CDA {num_cda} e CNPJ {icnpj}")

        driver.get("https://www.sefaz.rs.gov.br/DAT/DAT-GAU-EMI-DIV_1.aspx?cpf_fis=&cnpj_empresa=&cgcte_empresa=&ipva=")
        time.sleep(random.uniform(4, 6))

        logging.info(f"CNPJ recebido pela função: '{icnpj}'")

        try:
            cnpj = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/fieldset[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/input[1]')
            cnpj.clear()

            for char in str(icnpj):
                cnpj.send_keys(char)
                time.sleep(random.uniform(0.1, 0.2))

            time.sleep(1)

            abr = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/fieldset[2]/div[1]/table/tbody/tr/td/table/tbody/tr[1]/td[2]/input[2]')
            abr.click()

            nat_debito = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/fieldset[2]/div[1]/table/tbody/tr/td/table/tbody/tr[2]/td[2]/select')
            nat_debito.click()
            time.sleep(1)

            todos = driver.find_element(By.XPATH, '//*[text()="Todos"]')
            todos.click()
            time.sleep(1)

            quitacao = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/fieldset[2]/div[1]/table/tbody/tr/td/table/tbody/tr[3]/td[2]/input[2]')
            quitacao.click()
            
            data_penultimo_dia = penultimo_dia_util()
            data = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/fieldset[2]/div[1]/table/tbody/tr/td/table/tbody/tr[4]/td[2]/input')
            data.click()
            data.clear()

            for char in data_penultimo_dia:
                data.send_keys(char)
                time.sleep(random.uniform(0.1, 0.2))
            
            time.sleep(1)

            cda = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/fieldset[2]/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[2]/input')
            cda.click()
            cda.clear()

            for char in str(num_cda):
                cda.send_keys(char)
                time.sleep(random.uniform(0.2, 0.3))

            time.sleep(1)

            avancar = driver.find_element(By.XPATH, '/html/body/div[4]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div/fieldset/div/b/table/tbody/tr/td/input')
            avancar.click()
            time.sleep(2)

            checkbox = driver.find_element(By.XPATH, '/html/body/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div[1]/fieldset/div/fieldset[3]/table/tbody/tr[1]/th/input')
            if not checkbox.is_selected():
                checkbox.click()
                print("Checkbox marcado!")
            
            else:
                print("Checkbox já estava marcado, nenhuma ação necessária.")
            
            download = driver.find_element(By.XPATH, '/html/body/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/div/form/div[1]/fieldset/div/fieldset[4]/input')
            download.click()
            time.sleep(7)

            boleto = driver.find_element(By.XPATH, '/html/body/div[2]/table/tbody/tr[1]/td/table/tbody/tr/td[2]/div/div[2]/fieldset/div/fieldset[2]/table/tbody/tr[2]/td[6]/a[2]/img')
            arquivos_antes = set(os.listdir(pasta_download))
            logging.info(f"Arquivos antes do download: {len(arquivos_antes)}")
            boleto.click()
            logging.info("Fazendo download...")
            time.sleep(2)

            arquivo_encontrado = None
            max_espera = 15
            inicio = time.time()

            while time.time() - inicio < max_espera:
                arquivos_atuais = set(os.listdir(pasta_download))
                novos_arquivos = arquivos_atuais - arquivos_antes

                for arquivo in novos_arquivos:
                    if arquivo.startswith("GA_SEFAZRS_") and arquivo.endswith(".pdf"):
                        arquivo_encontrado = os.path.join(pasta_download, arquivo)
                        logging.info(f"Arquivo encontrado: {arquivo}")
                        break
                
                if arquivo_encontrado:
                    break

                if not novos_arquivos:
                    for arquivo in arquivos_atuais:
                        if arquivo.startswith("GA_SEFAZRS_") and arquivo.endswith(".pdf"):
                            caminho_completo = os.path.join(pasta_download, arquivo)
                            tempo_modificacao = os.path.getmtime(caminho_completo)
                            if time.time() - tempo_modificacao < 5:
                                arquivo_encontrado = caminho_completo
                                logging.info(f"Arquivo recentemente modificado encontrado: {arquivo}")
                                break
                
                if arquivo_encontrado:
                    break

                time.sleep(1)
            
            if arquivo_encontrado:
                nome_empresa_limpo = limpar_nome_arquivo(resultado["nome_empresa"])
                novo_nome = os.path.join(pasta_download, f"GA_SEFAZRS_{nome_empresa_limpo} - {num_cda}.pdf")

                if os.path.exists(novo_nome):
                    os.remove(novo_nome)
                
                shutil.move(arquivo_encontrado, novo_nome)
                logging.info(f"Arquivo renomeado de {os.path.basename(arquivo_encontrado)} para: GA_SEFAZRS_{nome_empresa_limpo} - {num_cda}.pdf")

                resultado["sucesso"] = True
                resultado["data_vencimento"] = f"15/{mes_atual_texto[:3]}/{ano_atual}"
                return resultado
            
            else:
                logging.error("Arquivo não foi baixado após o tempo de espera")
                resultado["sucesso"] = False
                resultado["motivo_falha"] = "Arquivo não foi baixado"
                return resultado
        
        except Exception as e:
            logging.error(f"Erro ao tentar preencher algum campo: {e}")
            resultado["sucesso"] = False
            resultado["motivo_falha"] = f"Erro ao preencher campo: {str(e)[:100]}"
            return resultado
    
    except Exception as e:
        logging.error(f"Erro ao tentar entrar no site: {e}")
        resultado["sucesso"] = False
        resultado["motivo_falha"] = str(e)[:100]
        return resultado

def pintar_planilha(caminho_planilha, resultado):
    try:
        logging.info(f"Iniciando processo de coloração da planilha: {caminho_planilha}")

        wb = openpyxl.load_workbook(caminho_planilha)

        data_hoje = datetime.now()
        mes_atual = data_hoje.month
        ano_atual = data_hoje.year

        meses = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
        }
        nome_aba = f"{meses[mes_atual]} {ano_atual}"

        if nome_aba not in wb.sheetnames:
            logging.error(f"Aba {nome_aba} não encontrada na planilha")
            return False
        
        ws = wb[nome_aba]

        coluna_situacao_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Situação":
                coluna_situacao_idx = col_idx
                break
        
        if coluna_situacao_idx is None:
            coluna_situacao_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_situacao_idx).value = "Situação"
            ws.cell(row=1, column=coluna_situacao_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Situação' criada na posição {coluna_situacao_idx}")

        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        azul = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        mes_atual_num = f"{mes_atual:02d}/{ano_atual}"

        num_parcelamento = resultado["num_parcelamento"]
        sucesso = resultado["sucesso"]
        data_vencimento = resultado["data_vencimento"]
        motivo_falha = resultado.get("motivo_falha", "")
        situacao_especial = resultado.get("situacao_especial", "")

        linha_encontrada = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[5].value).strip() == str(num_parcelamento).strip():
                linha_encontrada = row_idx
                break
            
        if not linha_encontrada:
            logging.warning(f"Linha para parcelamento {num_parcelamento} não encontrada na planilha")
            return False

        cor_a_aplicar = None
        texto_situacao = ""

        if sucesso:
            try:
                partes_data = data_vencimento.split('/')

                meses_str_para_num = {
                    'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
                    'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
                    'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'
                }
                mes_raw = partes_data[1].strip().upper()

                if mes_raw.isdigit():
                    mes_site = f"{int(mes_raw):02d}/{partes_data[2]}"
                
                else:
                    mes_abrev = mes_raw[:3]
                    mes_site = f"{meses_str_para_num.get(mes_abrev, '00')}/{partes_data[2]}"

                if mes_site == mes_atual_num:
                    cor_a_aplicar = verde
                    texto_situacao = "OK"
                    logging.info(f"Pintado linha {linha_encontrada} de VERDE (data correspondente)")
                    
                else:
                    cor_a_aplicar = amarelo
                    texto_situacao = f"OK / Data Divergente ({data_vencimento})"
                    logging.info(f"Pintando linha {linha_encontrada} de AMARELO (data diferente: {mes_site} vs {mes_atual_num})")
                
            except:
                cor_a_aplicar = amarelo
                texto_situacao = f"OK / Data Divergente (formato inválido: {data_vencimento})"
                logging.warning(f"Formato de data inesperado: {data_vencimento}, pintado de AMARELO")
            
        elif situacao_especial == "quitado":
            cor_a_aplicar = azul
            texto_situacao = "Parcelamento Quitado"
            logging.info(f"Pintando linha {linha_encontrada} de AZUL (parcelamento quitado)")
            
        elif situacao_especial == "desistente":
            cor_a_aplicar = vermelho
            texto_situacao = "Parcelamento Desistente"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (parcelamento desistente)")
            
        else:
            cor_a_aplicar = vermelho
            texto_situacao = f"ERRO: {motivo_falha}"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (falha no processamento)")

        cell = ws.cell(row=linha_encontrada, column=coluna_situacao_idx)
        cell.value = texto_situacao
        
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if linha_encontrada == 2:
            col_letter = get_column_letter(coluna_situacao_idx)
            ws.column_dimensions[col_letter].width = 40
        
        coluna_atrasadas_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Atrasadas?":
                coluna_atrasadas_idx = col_idx
                break
        
        if coluna_atrasadas_idx is None:
            coluna_atrasadas_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_atrasadas_idx).value = "Atrasadas?"
            ws.cell(row=1, column=coluna_atrasadas_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Atrasadas?' criada na posição {coluna_atrasadas_idx}")
        
        valor_atrasadas = resultado.get("atrasadas", "NÃO")
        ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).value = valor_atrasadas

        total_colunas = ws.max_column
        for col_idx in range(1, total_colunas + 1):
            ws.cell(row=linha_encontrada, column=col_idx).fill = cor_a_aplicar
        
        cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if valor_atrasadas.startswith("SIM"):
            ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).fill=cinza
        
        ajustar_largura_colunas(ws)

        wb.save(caminho_planilha)
        logging.info(f"Linha {linha_encontrada} colorida, situação atualizada e planilha salva com sucesso")
        return True
    
    except Exception as e:
        logging.error(f"Erro ao pintar linha planilha: {e}")
        logging.error(traceback.format_exc())
        return False
    
def main():

    data_hoje = datetime.now()
    meses_abrev = {1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR', 5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO', 9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'}
    mes_atual_abrev = meses_abrev[data_hoje.month]
    ano = data_hoje.year
    pasta_mensal = os.path.join(PASTA_BASE, f"{mes_atual_abrev} {ano}")

    try:
        driver, pasta_download = configurar_navegador()
        
        dados, coluna_tipo, coluna_numero, mes_atual_texto = carregar_dados_planilha_mensal()
        
        if not dados is None and not coluna_tipo is None and not coluna_numero is None:
            logging.info(f"Dados carregados com sucesso da planilha-mãe")           
            
            dados_filtrados = dados.dropna(subset=[coluna_tipo])
            dados_filtrados = dados_filtrados[dados_filtrados[coluna_tipo].str.lower().str.strip().str.contains(r'estadual\s*-?\s*rs', regex=True)]
            logging.info(f"Total de registros filtrados: {len(dados_filtrados)}")

            if len(dados_filtrados) == 0:
                logging.warning("Nenhum número de parcelamento 'ESTADUAL - RS' encontrado na planilha!")
                return
            
            dados_filtrados['Num_Parcelamento_Limpo'] = dados_filtrados[coluna_numero].apply(extrair_numero_parcelamento)
            dados_filtrados['CDA'] = dados_filtrados[coluna_numero].apply(extrair_numero_cda)

            coluna_cnpj = None
            for col in dados.columns:
                col_str = str(col)
                col_lower = col_str.lower()
                if "cnpj" in col_lower:
                    coluna_cnpj = col
                    logging.info(f"Coluna de CNPJ identificada: '{col}'")
                    break
            
            if coluna_cnpj is None:
                logging.warning("Coluna de CNPJ não encontrada na planilha!")
                coluna_cnpj = 'CNPJ FORMATADO'

            caminho_planilha = CAMINHO_PLANILHA

            for _, row in dados_filtrados.iterrows():
                try:
                    num_parcelamento_completo = row[coluna_numero]
                    num_parcelamento = row['Num_Parcelamento_Limpo']
                    num_cda = row['CDA']
                    nome_empresa = row['EMPRESA']

                    if pd.isna(num_cda):
                        logging.warning(f"CDA não encontrado para o parcelamento {num_parcelamento}. Pulando...")
                        continue

                    logging.info(f"Processando CDA: {num_cda} - Parcelamento: {num_parcelamento} - Empresa: {nome_empresa}")

                    icnpj = row.get(coluna_cnpj, '')
                    logging.info(f"Parcelamento: {num_parcelamento}, CNPJ bruto: {icnpj}")

                    if pd.isna(icnpj) or icnpj == '' or icnpj is None:
                        logging.warning(f"CNPJ não encontrado para o parcelamento {num_parcelamento}. Pulando este parcelamento.")

                        resultado = {
                            "sucesso": False,
                            "data_vencimento": None,
                            "motivo_falha": "CNPJ não encontrado na planilha.",
                            "num_parcelamento": num_parcelamento,
                            "nome_empresa": nome_empresa
                        }
                    
                        if pintar_planilha(caminho_planilha, resultado):
                            logging.info(f"Linha para parcelamento {num_parcelamento} colorida com sucesso")

                        else:
                            logging.error(f"Falha ao colorir linha para parcelamento {num_parcelamento}")
                    
                        continue

                    icnpj = limpar_cnpj(icnpj)
                    logging.info(f"CNPJ limpo: {icnpj}")

                    if len(icnpj) < 14:
                        icnpj = icnpj.zfill(14)
                        logging.info(f"CNPJ ajustado para 14 dígitos: {icnpj}")
            
                    if 'CNPJ FORMATADO' in row:
                        logging.info(f"Coluna 'CNPJ FORMATADO' encontrada com valor: {row['CNPJ FORMATADO']}")
                
                    else:
                        logging.warning(f"Coluna 'CNPJ FORMATADO' não encontrada nas colunas disponíveis: {list(row.index)}")
                    
                    resultado = {
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": None,
                        "num_parcelamento": num_parcelamento,
                        "nome_empresa": nome_empresa
                    }

                    resultado = baixar_dae_mes_atual(driver=driver, num_cda=num_cda, mes_atual_texto=mes_atual_texto, ano_atual=datetime.now().year, 
                                                     icnpj=icnpj, pasta_download=pasta_download, resultado=resultado)

                    if resultado["sucesso"]:
                        logging.info(f"Processamento bem-sucedido para parcelamento {num_parcelamento}")

                        mes_vencimento = resultado["data_vencimento"]
                        if mes_vencimento:
                            try:
                                venc = datetime.strptime(mes_vencimento[:10], "%d/%b/%Y")
                            except:
                                try:
                                    venc = datetime.strptime(mes_vencimento[:10], "%d/%m/%Y")
                                except:
                                    venc = None

                            if venc and venc < datetime.now():
                                resultado["atrasadas"] = "SIM"

                            else:
                                resultado["atrasadas"] = "NÃO"

                        else:
                            resultado["atrasadas"] = "NÃO"
                    
                    else:
                        logging.warning(f"Falha no processamento do parcelamento {num_parcelamento}: {resultado['motivo_falha']}")             

                    if pintar_planilha(caminho_planilha, resultado):
                        logging.info(f"Linha para parcelamento {num_parcelamento} colorida com sucesso")

                    else:
                        logging.error(f"Falha ao colorir linha para parcelamento {num_parcelamento}")

                    time.sleep(random.uniform(3,5))        
                
                except Exception as e:
                    logging.error(f"Erro ao processar linha para parcelamento: {e}")

                    resultado = {
                        "num_parcelamento": num_parcelamento if 'num_parcelamento' in locals() else "Desconhecido",
                        "nome_empresa": nome_empresa if 'nome_empresa' in locals() else "Desconhecido",
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": str(e)[:100]
                    }
                    pintar_planilha(caminho_planilha, resultado)
                    continue
            
            logging.info(f"Processamento concluído para {len(dados_filtrados)} parcelamentos")
        
        else:
            logging.error("Não foi possível carregar os dados da planilha-mãe")
    
    except Exception as e:
        logging.error(f"Erro no processo principal: {e}")
    
    finally:
        try:
            if 'driver' in locals() and driver is not None:
                try:
                    driver.quit()
                except Exception as e:
                    logging.warning(f"Erro ao encerrar WebDriver: {e}")
                finally:
                    driver = None
            logging.info("Navegador fechado")
        except Exception:
            pass

if __name__ == "__main__":
    try:
        logging.info("Iniciando script de automação de parcelamentos")
        main()
        logging.info("Script finalizado com sucesso")
    
    except Exception as e:
        logging.error(f"Erro fatal: {e}")
    
    print("\nProcessamento concluído!")
    sys.exit(0)
