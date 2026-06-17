import undetected_chromedriver as uc
import re
import pandas as pd
import sys
import time
import os
import random
import openpyxl
import logging
import shutil
import traceback
import math
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from twocaptcha import TwoCaptcha
from config import TWOCAPTCHA_API_KEY
from pathlib import Path
from utils import (
    carregar_dados_planilha_mensal,
    ajustar_largura_colunas,
    CAMINHO_PLANILHA,
    PASTA_DOWNLOADS,
    PASTA_BASE,
    MESES_ABREV,
    MESES_NUMERO,
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)

def configurar_navegador():
    pasta_download_base = PASTA_BASE
    subpasta = "Parcelamentos - MG"
    pasta_download = criar_estrutura_pastas_por_vencimento(pasta_download_base, subpasta)

    logging.info(f"Usando pasta de download: {pasta_download}")

    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")

    prefs = {
        "download.default_directory": pasta_download,
        "download.prompt_for_download": False,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "profile.default_content_setting_values.popups": 1

    }
    options.add_experimental_option("prefs", prefs)

    driver = uc.Chrome(options=options)

    params = {
        "behavior": "allow",
        "downloadPath": pasta_download
    }
    driver.execute_cdp_cmd("Page.setDownloadBehavior", params)

    return driver, pasta_download

def resolver_captcha(driver):
    max_tentativas = 3
    
    for tentativa in range(1, max_tentativas + 1):
        try:
            print(f"[INFO] Tentativa {tentativa}/{max_tentativas}")
            
            solver = TwoCaptcha(TWOCAPTCHA_API_KEY)
            
            # Aguardar elemento Turnstile carregar completamente
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, "cf-turnstile"))
            )
            time.sleep(2)  # Tempo adicional para estabilizar
            
            turnstile_div = driver.find_element(By.CLASS_NAME, "cf-turnstile")
            sitekey = turnstile_div.get_attribute("data-sitekey")
            url = driver.current_url
            
            # Validar sitekey
            if not sitekey or len(sitekey) < 10:
                raise Exception(f"Sitekey inválida: {sitekey}")
            
            print(f"[INFO] Sitekey encontrado: {sitekey}")
            print(f"[INFO] URL da página: {url}")
            
            # Resolver captcha com timeout estendido
            result = solver.turnstile(sitekey=sitekey, url=url, timeout=180)
            token = result['code']
            
            print(f"[INFO] Token recebido: {token[:20]}...")
            
            # Injetar token com múltiplas tentativas de seletor
            injection_scripts = [
                f'document.querySelector("input[name=\'cf-turnstile-response\']").value = "{token}";',
                f'document.querySelector("textarea[name=\'cf-turnstile-response\']").value = "{token}";',
                f'document.querySelector("[name=\'cf-turnstile-response\']").value = "{token}";'
            ]
            
            token_injetado = False
            for script in injection_scripts:
                try:
                    driver.execute_script(script)
                    token_injetado = True
                    print("[INFO] Token injetado com sucesso")
                    break
                except:
                    continue
            
            if not token_injetado:
                raise Exception("Falha ao injetar token - elemento não encontrado")
            
            time.sleep(3)
            print("[INFO] ✅ Captcha resolvido pelo 2Captcha!")
            return True
            
        except Exception as e:
            error_msg = str(e).lower()
            
            # Tratamento específico por tipo de erro
            if "sitekey" in error_msg or "invalid" in error_msg:
                print(f"[ERRO] Tentativa {tentativa} - Sitekey inválida: {str(e)}")
            elif "timeout" in error_msg or "time" in error_msg:
                print(f"[ERRO] Tentativa {tentativa} - Timeout na resolução: {str(e)}")
            elif "balance" in error_msg or "insufficient" in error_msg:
                print(f"[ERRO] Tentativa {tentativa} - Problema de saldo: {str(e)}")
            elif "element" in error_msg or "not found" in error_msg:
                print(f"[ERRO] Tentativa {tentativa} - Elemento não encontrado: {str(e)}")
            elif "network" in error_msg or "connection" in error_msg:
                print(f"[ERRO] Tentativa {tentativa} - Problema de conexão: {str(e)}")
            else:
                print(f"[ERRO] Tentativa {tentativa} - Falha geral: {str(e)}")
            
            # Se não é a última tentativa, aguardar antes de tentar novamente
            if tentativa < max_tentativas:
                print(f"[INFO] Aguardando 5 segundos antes da próxima tentativa...")
                time.sleep(5)
            else:
                print(f"[ERRO] ❌ Falha ao resolver Turnstile após {max_tentativas} tentativas")
    
    return False

def limpar_nome_arquivo(nome):
    caracteres_invalidos = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']

    nome_limpo = nome
    for char in caracteres_invalidos:
        nome_limpo = nome_limpo.replace(char, '_')
    
    if len(nome_limpo) > 50:
        nome_limpo = nome_limpo[:47] + '...'
    
    return nome_limpo.strip()

def criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, data_vencimento=None):
    meses_abrev = {
        1: 'JAN', 2: 'FEV', 3: 'MAR', 4: 'ABR',
        5: 'MAI', 6: 'JUN', 7: 'JUL', 8: 'AGO',
        9: 'SET', 10: 'OUT', 11: 'NOV', 12: 'DEZ'
    }
    
    if data_vencimento:
        try:
            partes_data = data_vencimento.split('/')
            
            mes_raw = partes_data[1].strip().upper()
            ano = partes_data[2].strip()
            
            if mes_raw.isdigit():
                mes_num = int(mes_raw)

            else:
                meses_str_para_num = {
                    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4,
                    'MAI': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
                    'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12
                }
                mes_abrev = mes_raw[:3]
                mes_num = meses_str_para_num.get(mes_abrev, datetime.now().month)
            
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
            
        except Exception as e:
            logging.warning(f"Erro ao processar data de vencimento '{data_vencimento}', usando mês atual: {str(e)}")
            data_atual = datetime.now()
            mes_num = data_atual.month
            ano = str(data_atual.year)
            mes_abrev_pasta = meses_abrev[mes_num]
            meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
            numero_mes = meses_dict[mes_abrev_pasta.upper()]
            nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"

    else:
        data_atual = datetime.now()
        mes_num = data_atual.month
        ano = str(data_atual.year)
        mes_abrev_pasta = meses_abrev[mes_num]
        meses_dict = {
                "JAN": "01", "FEV": "02", "MAR": "03", "ABR": "04",
                "MAI": "05", "JUN": "06", "JUL": "07", "AGO": "08",
                "SET": "09", "OUT": "10", "NOV": "11", "DEZ": "12"
            }
        numero_mes = meses_dict[mes_abrev_pasta.upper()]
        nome_pasta_mes = f"{numero_mes} - {mes_abrev_pasta} {ano}"
    
    caminho_pasta_mes = os.path.join(pasta_base, nome_pasta_mes)
    
    if not os.path.exists(caminho_pasta_mes):
        os.makedirs(caminho_pasta_mes)
        logging.info(f"Pasta do mês criada: {caminho_pasta_mes}")
    
    caminho_completo = os.path.join(caminho_pasta_mes, subpasta)
    
    if not os.path.exists(caminho_completo):
        os.makedirs(caminho_completo)
        logging.info(f"Subpasta criada: {caminho_completo}")
    
    return caminho_completo

def baixar_dae_mes_atual(driver, num_parcelamento, mes_atual_texto, ano_atual, pasta_download, nome_empresa):
    try:
        logging.info(f"Iniciando processo para parcelamento: {num_parcelamento}")

        resultado = {
            "sucesso": False,
            "data_vencimento": None,
            "motivo_falha": None
        }
        
        driver.get("https://www2.fazenda.mg.gov.br/sol/ctrl/SOL/PARCEL/CONSULTA_003?ACAO=VISUALIZAR#")
        time.sleep(random.uniform(2, 5))

        try:
            tipo_parcelamento = WebDriverWait(driver, 25).until(
                EC.element_to_be_clickable((By.XPATH, 
                "/html/body/div[3]/div[2]/div/div[3]/div/form/table[3]/tbody/tr[1]/td[2]/input")))
            
            actions = ActionChains(driver)
            (actions.move_to_element(tipo_parcelamento)
             .pause(random.uniform(0.3, 1.2))
             .click()
             .perform())
            
            time.sleep(random.uniform(0.8, 1.5))
            logging.info("Tipo de consulta selecionado")

        except Exception as e:
            logging.error(f"Falha ao selecionar tipo: {str(e)[:100]}...")
            resultado["motivo_falha"] = f"Falha ao selecionar tipo: {str(e)[:100]}"
            return resultado

        try:
            input_parcelamento = WebDriverWait(driver, 25).until(
                EC.element_to_be_clickable((By.XPATH, 
                "/html/body/div[3]/div[2]/div/div[3]/div/form/div[3]/table/tbody/tr[1]/td[2]/input")))

            input_parcelamento.clear()
            for char in str(num_parcelamento):
                input_parcelamento.send_keys(char)
                time.sleep(random.uniform(0.1, 0.3))
                
            time.sleep(random.uniform(0.5, 1.5))
            logging.info("Número preenchido humanamente")

        except Exception as e:
            logging.error(f"Falha ao preencher número: {str(e)[:100]}...")
            resultado["motivo_falha"] = f"Falha ao preencher número: {str(e)[:100]}"
            return resultado

        max_tentativas = 2
        for tentativa in range(1, max_tentativas+1):
            if resolver_captcha(driver):
                logging.info("Captcha resolvido com sucesso pelo 2captcha")
                break
                
            if tentativa < max_tentativas:
                logging.warning(f"Tentativa {tentativa} falhou, tentando novamente...")
                time.sleep(3)

            else:
                logging.warning("Falha no captcha automático, solicitando manual...")
                print("\n" + "="*50)
                print("RESOLVA O CAPTCHA MANUALMENTE:")
                print(f"1. Abra a janela do navegador")
                print(f"2. Resolva o captcha (se aparecer)")
                print(f"3. Volte aqui e pressione ENTER")
                print("="*50 + "\n")
                input("Pressione ENTER após resolver...")
                logging.info("Captcha manual resolvido")

        max_tentativas_click = 3
        for tentativa in range(1, max_tentativas_click + 1):
            try:
                avancar = WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.XPATH, 
                        "/html/body/div[3]/div[2]/div/div[3]/div/form/table[4]/tbody/tr[2]/td/a")))
                try:
                    avancar.click()

                except:
                    try:
                        driver.execute_script("arguments[0].click();", avancar)

                    except:
                        actions = ActionChains(driver)
                        actions.move_to_element(avancar).pause(0.5).click().perform()
                
                logging.info(f"Clicando em avancar na tentativa {tentativa}")
                break

            except Exception as e:
                if tentativa == max_tentativas_click:
                    logging.error(f"Falha ao clicar em avançar após {max_tentativas_click} tentativas: {e}")
                    resultado["motivo_falha"] = f"Falha ao clicar em avançar: {str(e)[:100]}"
                    return resultado
                else:
                    logging.warning(f"Tentativa {tentativa} falhou, tentando novamente...")
                    time.sleep(1)

        try:
            situacao_element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH,
                    '/html/body/div[3]/div[2]/div/div[3]/div/form/table[3]/tbody/tr[3]/td[2]/span/span'))
            )
            situacao_text = situacao_element.text.strip()
            logging.info(f"Situação do parcelamento: {situacao_text}")
            
            if "desistente" in situacao_text.lower():
                logging.info(f"Parcelamento {num_parcelamento} está na situação 'Desistente'. Finalizando processo para este parcelamento.")
                resultado["motivo_falha"] = "Parcelamento na situação 'Desistente'"
                resultado["situacao_especial"] = "desistente"
                return resultado
            
            elif "quitado" in situacao_text.lower():
                logging.info(f"Parcelamento {num_parcelamento} está na situação 'Quitado'. Finalizando processo para este parcelamento.")
                resultado["motivo_falha"] = "Parcelamento na situação 'Quitado'"
                resultado["situacao_especial"] = "quitado"
                return resultado
        
        except Exception as e:
            logging.error(f"Falha ao verificar situação do parcelamento: {e}")

        guia_encontrada = False

        try:
            time.sleep(3)

            qtd_parcelas_elem = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div[2]/div/div[3]/div/form/table[3]/tbody/tr[4]/td[2]'))
            )
            qtd_parcelas = int(qtd_parcelas_elem.text.strip())
            total_paginas = math.ceil(qtd_parcelas / 11)

            logging.info(f"Quantidade de parcelas: {qtd_parcelas} — Estimando {total_paginas} páginas.")
        
        except Exception as e:
            total_paginas = 1
            logging.warning(f"Erro ao determinar número de páginas, assumindo 1. Erro: {str(e)[:100]}")

        mes_atual_num = f"{datetime.now().month:02d}/{datetime.now().year}"
        pagina_atual = 1
        guia_encontrada = False
        guias_atrasadas = []
        primeira_guia_nao_quitada = None
        
        while not guia_encontrada and pagina_atual <= total_paginas:
            logging.info(f"Verificando página {pagina_atual} de {total_paginas}")
            
            if pagina_atual > 1:
                try:
                    campo_pagina = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@name='ufw_posicionador_gridParcelas']"))
                    )
                    campo_pagina.clear()
                    campo_pagina.send_keys(str(pagina_atual))
                    time.sleep(1)

                    botao_ir = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@name='ufw_link_ir_gridParcelas']"))
                    )
                    botao_ir.click()

                    logging.info(f"Navegando diretamente para a página {pagina_atual}")

                    WebDriverWait(driver, 8).until(
                        lambda d: f"{pagina_atual} de" in d.find_element(By.XPATH, "//td[contains(@class, 'ctnhdr')]").text
                    )
                
                except Exception as e:
                    logging.warning(f"Erro ao navegar para a página {pagina_atual}: {str(e)[:100]}")
            
            encontrou_linha = False
            for linha_atual in range(1, 13):
                try:
                    xpath_linha = f'/html/body/div[3]/div[2]/div/div[3]/div/form/table[7]/tbody/tr[{linha_atual}]'
                    linha = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located((By.XPATH, xpath_linha))
                    )
                    
                    encontrou_linha = True
                    texto_linha = linha.text
                    logging.info(f"Página {pagina_atual}, Linha {linha_atual}: {texto_linha}")
                    
                    if "NAO QUITADA" in texto_linha.upper():
                        xpath_data = f'/html/body/div[3]/div[2]/div/div[3]/div/form/table[7]/tbody/tr[{linha_atual}]/td[3]'
                        data_elemento = driver.find_element(By.XPATH, xpath_data)
                        data_vencimento = data_elemento.text.strip()
                        
                        try:
                            partes_data = data_vencimento.split('/')
                            meses_str_para_num = {
                                'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
                                'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
                                'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'
                            }
                            
                            mes_raw = partes_data[1].strip().upper()

                            if mes_raw.isdigit():
                                mes_site = f"{int(mes_raw):02d}/{partes_data[2]}"

                            else:
                                mes_abrev = mes_raw[:3]
                                mes_site = f"{meses_str_para_num.get(mes_abrev, '00')}/{partes_data[2]}"
                            
                            if mes_site == mes_atual_num:
                                logging.info(f"Encontrada parcela NÃO QUITADA do MÊS ATUAL com vencimento em: {data_vencimento}")

                                xpath_dae = f"/html/body/div[3]/div[2]/div/div[3]/div/form/table[7]/tbody/tr[{linha_atual}]/td[4]/a"
                                botao_dae = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((By.XPATH, xpath_dae))
                                )
                                botao_dae.click()

                                time.sleep(4)

                                arquivo_original = os.path.join(pasta_download, "SERVICO_002.pdf")
                                max_espera = 5
                                inicio = time.time()

                                while time.time() - inicio < max_espera:
                                    if os.path.exists(arquivo_original):
                                        tamanho_inicial = os.path.getsize(arquivo_original)
                                        time.sleep(1)
                                        if os.path.getsize(arquivo_original) == tamanho_inicial:
                                            break
                                    time.sleep(1)
                                
                                if os.path.exists(arquivo_original):
                                    pasta_base = PASTA_BASE
                                    subpasta = "Parcelamentos - MG"
                                    pasta_destino = criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, data_vencimento)

                                    nome_empresa_limpo = limpar_nome_arquivo(nome_empresa)
                                    novo_nome = os.path.join(pasta_destino, f"DAE_{nome_empresa_limpo} - {num_parcelamento}.pdf")

                                    if os.path.exists(novo_nome):
                                        os.remove(novo_nome)
                                    
                                    shutil.move(arquivo_original, novo_nome)
                                    logging.info(f"Arquivo renomeado e movido para: {novo_nome}")
                                
                                else:
                                    logging.warning(f"Arquivo de download não encontrado após {max_espera} segundos")
                                    resultado["motivo_falha"] = "Arquivo de download não encontrado"

                                resultado["data_vencimento"] = data_vencimento
                                resultado["sucesso"] = True
                                guia_encontrada = True
                                break

                            else:
                                # NOVA LÓGICA: Verificar se é atrasada OU adiantada
                                try:
                                    data_venc_obj = datetime.strptime(data_vencimento, "%d/%m/%Y")
                                    data_hoje = datetime.now()
                                    
                                    if data_venc_obj < data_hoje:
                                        # É atrasada
                                        guias_atrasadas.append({
                                            "linha": linha_atual,
                                            "pagina": pagina_atual,
                                            "data": data_vencimento,
                                            "mes_site": mes_site,
                                            "tipo": "atrasada"
                                        })
                                        logging.info(f"Guia ATRASADA encontrada: {data_vencimento}")
                                    elif data_venc_obj > data_hoje:
                                        # É adiantada - armazenar a primeira encontrada
                                        if primeira_guia_nao_quitada is None:
                                            primeira_guia_nao_quitada = {
                                                "linha": linha_atual,
                                                "pagina": pagina_atual,
                                                "data": data_vencimento,
                                                "mes_site": mes_site,
                                                "tipo": "adiantada"
                                            }
                                            logging.info(f"Primeira guia ADIANTADA encontrada: {data_vencimento}")
                                except:
                                    # Se não conseguir parsear a data, considera como atrasada
                                    guias_atrasadas.append({
                                        "linha": linha_atual,
                                        "pagina": pagina_atual,
                                        "data": data_vencimento,
                                        "mes_site": mes_site,
                                        "tipo": "atrasada"
                                    })
                                    logging.info(f"Guia com data não parseável considerada ATRASADA: {data_vencimento}")
                        
                        except Exception as e:
                            logging.warning(f"Erro ao processar data {data_vencimento}: {str(e)[:100]}")

                            if primeira_guia_nao_quitada is None:
                                primeira_guia_nao_quitada = {
                                    "linha": linha_atual,
                                    "pagina": pagina_atual,
                                    "data": data_vencimento,
                                    "mes_site": "erro"
                                }

                except Exception as e:
                    logging.info(f"Linha {linha_atual} não encontrada ou erro ao processar: {str(e)[:100]}")
                    continue
            
            if guia_encontrada:
                break
                
            if not encontrou_linha:
                logging.warning(f"Nenhuma linha encontrada na página {pagina_atual}")
            
            pagina_atual += 1
            time.sleep(1)
        
        if not guia_encontrada and guias_atrasadas:
            logging.info(f"Encontradas {len(guias_atrasadas)} guias atrasadas. Iniciando downloads...")
            
            for idx, guia in enumerate(guias_atrasadas):
                try:
                    logging.info(f"Baixando guia atrasada {idx+1}/{len(guias_atrasadas)}: {guia['data']}")
                    
                    if pagina_atual != guia["pagina"]:
                        try:
                            campo_pagina = WebDriverWait(driver, 5).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@name='ufw_posicionador_gridParcelas']"))
                            )
                            campo_pagina.clear()
                            campo_pagina.send_keys(str(guia["pagina"]))
                            time.sleep(1)
                            botao_ir = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.XPATH, "//a[@name='ufw_link_ir_gridParcelas']"))
                            )
                            botao_ir.click()
                            time.sleep(2)
                            pagina_atual = guia["pagina"]
                        except Exception as e:
                            logging.warning(f"Erro ao navegar para página {guia['pagina']} (continuando mesmo assim): {str(e)[:100]}")
                    
                    linha_atual = guia["linha"]
                    xpath_dae = f"/html/body/div[3]/div[2]/div/div[3]/div/form/table[7]/tbody/tr[{linha_atual}]/td[4]/a"
                    
                    try:
                        botao_dae = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, xpath_dae))
                        )
                        botao_dae.click()
                        time.sleep(4)
                        
                        arquivo_original = os.path.join(pasta_download, "SERVICO_002.pdf")
                        max_espera = 8
                        inicio = time.time()
                        
                        while time.time() - inicio < max_espera:
                            if os.path.exists(arquivo_original):
                                tamanho_inicial = os.path.getsize(arquivo_original)
                                time.sleep(1)
                                if os.path.getsize(arquivo_original) == tamanho_inicial:
                                    break
                            time.sleep(1)
                        
                        if os.path.exists(arquivo_original):
                            pasta_base = PASTA_BASE
                            subpasta = "Parcelamentos - MG"
                            pasta_destino = criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, guia["data"])
                            
                            nome_empresa_limpo = limpar_nome_arquivo(nome_empresa)
                            novo_nome = os.path.join(pasta_destino, f"DAE_{nome_empresa_limpo} - {num_parcelamento} - {guia['data'].replace('/', '-')}.pdf")
                            
                            if os.path.exists(novo_nome):
                                os.remove(novo_nome)
                            
                            shutil.move(arquivo_original, novo_nome)
                            logging.info(f"Guia atrasada salva: {novo_nome}")
                            
                            if idx == 0:
                                resultado["data_vencimento"] = guia["data"]
                                resultado["sucesso"] = True
                                resultado["arquivo_salvo"] = novo_nome
                                
                        else:
                            logging.warning(f"Arquivo não encontrado para guia atrasada {guia['data']}")
                            
                    except Exception as e:
                        logging.error(f"Erro ao clicar no botão DAE da guia {guia['data']}: {str(e)[:100]}")
                        continue
                        
                except Exception as e:
                    logging.error(f"Erro ao processar guia atrasada {guia['data']}: {str(e)[:100]}")
                    continue
            
            if guias_atrasadas:
                datas_atrasadas = [g["data"] for g in guias_atrasadas]
                if len(datas_atrasadas) <= 3:
                    resultado["atrasadas"] = f"SIM - {', '.join(datas_atrasadas)}"
                else:
                    resultado["atrasadas"] = f"SIM - {', '.join(datas_atrasadas[:3])} (+{len(datas_atrasadas)-3} mais)"
                
                resultado["motivo_falha"] = f"Guia do mês atual não encontrada. Baixadas {len(guias_atrasadas)} guias atrasadas."
                
                if not resultado.get("sucesso"):
                    resultado["sucesso"] = True
                    resultado["data_vencimento"] = datas_atrasadas[0] if datas_atrasadas else "N/A"
        
        elif not guia_encontrada and not guias_atrasadas and primeira_guia_nao_quitada is not None:
            # Baixar a primeira guia adiantada
            logging.info(f"Nenhuma guia do mês atual ou atrasada encontrada. Baixando primeira guia adiantada: {primeira_guia_nao_quitada['data']}")
            
            try:
                # Navegar para a página da guia adiantada
                if pagina_atual != primeira_guia_nao_quitada["pagina"]:
                    try:
                        campo_pagina = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//input[@name='ufw_posicionador_gridParcelas']"))
                        )
                        campo_pagina.clear()
                        campo_pagina.send_keys(str(primeira_guia_nao_quitada["pagina"]))
                        time.sleep(1)
                        botao_ir = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((By.XPATH, "//a[@name='ufw_link_ir_gridParcelas']"))
                        )
                        botao_ir.click()
                        time.sleep(2)
                    except Exception as e:
                        logging.warning(f"Erro ao navegar para página da guia adiantada (continuando): {str(e)[:100]}")
                
                # Clicar no botão DAE da guia adiantada
                linha_atual = primeira_guia_nao_quitada["linha"]
                xpath_dae = f"/html/body/div[3]/div[2]/div/div[3]/div/form/table[7]/tbody/tr[{linha_atual}]/td[4]/a"
                
                botao_dae = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.XPATH, xpath_dae))
                )
                botao_dae.click()
                time.sleep(4)
                
                # Aguardar download
                arquivo_original = os.path.join(pasta_download, "SERVICO_002.pdf")
                max_espera = 8
                inicio = time.time()
                
                while time.time() - inicio < max_espera:
                    if os.path.exists(arquivo_original):
                        tamanho_inicial = os.path.getsize(arquivo_original)
                        time.sleep(1)
                        if os.path.getsize(arquivo_original) == tamanho_inicial:
                            break
                    time.sleep(1)
                
                if os.path.exists(arquivo_original):
                    # Salvar arquivo
                    pasta_base = PASTA_BASE
                    subpasta = "Parcelamentos - MG"
                    pasta_destino = criar_estrutura_pastas_por_vencimento(pasta_base, subpasta, primeira_guia_nao_quitada["data"])
                    
                    nome_empresa_limpo = limpar_nome_arquivo(nome_empresa)
                    novo_nome = os.path.join(pasta_destino, f"DAE_{nome_empresa_limpo} - {num_parcelamento} - {primeira_guia_nao_quitada['data'].replace('/', '-')}.pdf")
                    
                    if os.path.exists(novo_nome):
                        os.remove(novo_nome)
                    
                    shutil.move(arquivo_original, novo_nome)
                    logging.info(f"Guia adiantada salva: {novo_nome}")
                    
                    resultado["data_vencimento"] = primeira_guia_nao_quitada["data"]
                    resultado["sucesso"] = True
                    resultado["arquivo_salvo"] = novo_nome
                    resultado["atrasadas"] = "NÃO"
                    resultado["motivo_falha"] = f"Guia do mês atual não encontrada. Baixada guia adiantada de {primeira_guia_nao_quitada['data']}"
                else:
                    logging.warning("Arquivo de guia adiantada não encontrado")
                    
            except Exception as e:
                logging.error(f"Erro ao baixar guia adiantada: {str(e)[:100]}")
                resultado["motivo_falha"] = f"Erro ao baixar guia adiantada: {str(e)[:100]}"

        elif not guia_encontrada and not guias_atrasadas:
            logging.warning("Nenhuma guia não quitada foi encontrada.")
            resultado["motivo_falha"] = "Nenhuma guia não quitada encontrada"
            resultado["atrasadas"] = "NAO"

        if not guia_encontrada and primeira_guia_nao_quitada is None:
            logging.warning(f"Não foi encontrada nenhuma DAE NÃO QUITADA após verificar {pagina_atual-1} páginas")
            resultado["motivo_falha"] = "Nenhuma parcela NÃO QUITADA encontrada"

        if primeira_guia_nao_quitada:
            try:
                data_atraso = datetime.strptime(primeira_guia_nao_quitada['data'], "%d/%m/%Y")
                data_atual = datetime.now()

                if data_atraso.year < data_atual.year or (data_atraso.year == data_atual.year and data_atraso.month < data_atual.month):
                    resultado["atrasadas"] = f"SIM - {primeira_guia_nao_quitada['data']}"

                else:
                    resultado["atrasadas"] = "NÃO"
            
            except:
                resultado["atrasadas"] = "NÃO"
        
        else:
            resultado["atrasadas"] = "NÃO"  
        
        return resultado
        
    except Exception as e:
        logging.error(f"Erro no processo de baixar DAE: {e}")
        resultado["motivo_falha"] = str(e)[:100]
        return resultado

def pintar_planilha(caminho_planilha, resultado):
    try:
        logging.info(f"Iniciando processo de coloração da planilha: {caminho_planilha}")

        wb = openpyxl.load_workbook(caminho_planilha)

        data_hoje = datetime.now()
        mes_atual = data_hoje.month
        ano_atual = data_hoje.year

        meses = {
            1: "JANEIRO", 2: "FEVEREIRO", 3: "MARÇO", 4: "ABRIL",
            5: "MAIO", 6: "JUNHO", 7: "JULHO", 8: "AGOSTO",
            9: "SETEMBRO", 10: "OUTUBRO", 11: "NOVEMBRO", 12: "DEZEMBRO"
        }
        nome_aba = f"{meses[mes_atual]} {ano_atual}"

        if nome_aba not in wb.sheetnames:
            logging.error(f"Aba {nome_aba} não encontrada na planilha")
            return False
        
        ws = wb[nome_aba]

        coluna_situacao_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Situação":
                coluna_situacao_idx = col_idx
                break
        
        if coluna_situacao_idx is None:
            coluna_situacao_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_situacao_idx).value = "Situação"
            ws.cell(row=1, column=coluna_situacao_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Situação' criada na posição {coluna_situacao_idx}")

        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        azul = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        mes_atual_num = f"{mes_atual:02d}/{ano_atual}"

        num_parcelamento = resultado["num_parcelamento"]
        sucesso = resultado["sucesso"]
        data_vencimento = resultado["data_vencimento"]
        motivo_falha = resultado.get("motivo_falha", "")
        situacao_especial = resultado.get("situacao_especial", "")

        linha_encontrada = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[5].value).strip() == str(num_parcelamento).strip():
                linha_encontrada = row_idx
                break
            
        if not linha_encontrada:
            logging.warning(f"Linha para parcelamento {num_parcelamento} não encontrada na planilha")
            return False

        cor_a_aplicar = None
        texto_situacao = ""

        if sucesso:
            # Verificar se tem atrasadas baixadas
            valor_atrasadas = resultado.get("atrasadas", "NÃO")
            
            if valor_atrasadas.startswith("SIM"):
                # Se baixou atrasadas, pintar de verde mas manter informação das atrasadas
                cor_a_aplicar = verde
                texto_situacao = "OK"
                logging.info(f"Pintando linha {linha_encontrada} de VERDE (guias atrasadas baixadas)")
                
            else:
                # Lógica normal para verificar se é do mês atual
                try:
                    partes_data = data_vencimento.split('/')
                    meses_str_para_num = {
                        'JAN': '01', 'FEV': '02', 'MAR': '03', 'ABR': '04',
                        'MAI': '05', 'JUN': '06', 'JUL': '07', 'AGO': '08',
                        'SET': '09', 'OUT': '10', 'NOV': '11', 'DEZ': '12'
                    }
                    
                    mes_raw = partes_data[1].strip().upper()
                    if mes_raw.isdigit():
                        mes_site = f"{int(mes_raw):02d}/{partes_data[2]}"
                    else:
                        mes_abrev = mes_raw[:3]
                        mes_site = f"{meses_str_para_num.get(mes_abrev, '00')}/{partes_data[2]}"
                    
                    if mes_site == mes_atual_num:
                        cor_a_aplicar = verde
                        texto_situacao = "OK"
                        logging.info(f"Pintando linha {linha_encontrada} de VERDE (data correspondente)")
                    else:
                        cor_a_aplicar = amarelo
                        texto_situacao = f"OK / Data Divergente ({data_vencimento})"
                        logging.info(f"Pintando linha {linha_encontrada} de AMARELO (data adiantada)")
                except:
                    cor_a_aplicar = amarelo
                    texto_situacao = f"OK / Data Divergente (formato inválido: {data_vencimento})"
            
        elif situacao_especial == "quitado":
            cor_a_aplicar = azul
            texto_situacao = "Parcelamento Quitado"
            logging.info(f"Pintando linha {linha_encontrada} de AZUL (parcelamento quitado)")
            
        elif situacao_especial == "desistente":
            cor_a_aplicar = vermelho
            texto_situacao = "Parcelamento Desistente"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (parcelamento desistente)")
            
        else:
            cor_a_aplicar = vermelho
            texto_situacao = f"ERRO: {motivo_falha}"
            logging.info(f"Pintando linha {linha_encontrada} de VERMELHO (falha no processamento)")

        cell = ws.cell(row=linha_encontrada, column=coluna_situacao_idx)
        cell.value = texto_situacao
        
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if linha_encontrada == 2:
            col_letter = get_column_letter(coluna_situacao_idx)
            ws.column_dimensions[col_letter].width = 40
        
        coluna_atrasadas_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "Atrasadas?":
                coluna_atrasadas_idx = col_idx
                break
        
        if coluna_atrasadas_idx is None:
            coluna_atrasadas_idx = ws.max_column + 1
            ws.cell(row=1, column=coluna_atrasadas_idx).value = "Atrasadas?"
            ws.cell(row=1, column=coluna_atrasadas_idx).font = Font(bold=True)
            logging.info(f"Coluna 'Atrasadas?' criada na posição {coluna_atrasadas_idx}")
        
        valor_atrasadas = resultado.get("atrasadas", "NÃO")
        ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).value = valor_atrasadas

        total_colunas = ws.max_column
        for col_idx in range(1, total_colunas + 1):
            ws.cell(row=linha_encontrada, column=col_idx).fill = cor_a_aplicar
        
        cinza = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        if valor_atrasadas.startswith("SIM"):
            ws.cell(row=linha_encontrada, column=coluna_atrasadas_idx).fill=cinza

        ajustar_largura_colunas(ws)

        wb.save(caminho_planilha)
        logging.info(f"Linha {linha_encontrada} colorida, situação atualizada e planilha salva com sucesso")
        return True
    
    except Exception as e:
        logging.error(f"Erro ao pintar linha planilha: {e}")
        logging.error(traceback.format_exc())
        return False
    
def main():
    try:
        driver, pasta_download = configurar_navegador()
        
        dados, coluna_tipo, coluna_numero, mes_atual_texto = carregar_dados_planilha_mensal()
        
        if not dados is None and not coluna_tipo is None and not coluna_numero is None:
            logging.info(f"Dados carregados com sucesso da planilha-mãe")           
            
            dados_filtrados = dados.dropna(subset=[coluna_tipo])
            dados_filtrados = dados_filtrados[dados_filtrados[coluna_tipo].str.lower().str.strip().str.contains(r'estadual\s*-?\s*mg', regex=True)]
            logging.info(f"Total de registros filtrados: {len(dados_filtrados)}")

            if len(dados_filtrados) == 0:
                logging.warning("Nenhum número de parcelamento 'ESTADUAL - MG' encontrado na planilha!")
                return
            
            numeros_parcelamento = dados_filtrados[coluna_numero].dropna().unique().tolist()
            logging.info(f"Quantidade de parcelamentos encontrados: {len(numeros_parcelamento)}")
            logging.info(f"Números de parcelamentos encontrados: {numeros_parcelamento}")

            caminho_planilha = CAMINHO_PLANILHA

            for _, row in dados_filtrados.iterrows():
                try:
                    num_parcelamento = row[coluna_numero]
                    nome_empresa = row['EMPRESA']

                    logging.info(f"Processando parcelamento: {num_parcelamento} - Empresa: {nome_empresa}")

                    resultado = baixar_dae_mes_atual(driver, num_parcelamento, mes_atual_texto, datetime.now().year, pasta_download, nome_empresa)

                    resultado["num_parcelamento"] = num_parcelamento
                    resultado["nome_empresa"] = nome_empresa

                    if resultado["sucesso"]:
                        logging.info(f"Processamento bem-sucedido para parcelamento {num_parcelamento}")
                    
                    else:
                        logging.warning(f"Falha no processamento do parcelamento {num_parcelamento}: {resultado['motivo_falha']}")

                    if pintar_planilha(caminho_planilha, resultado):
                        logging.info(f"Linha para parcelamento {num_parcelamento} colorida com sucesso")

                    else:
                        logging.error(f"Falha ao colorir linha para parcelamento {num_parcelamento}")                     

                    time.sleep(random.uniform(3,5))
                
                except Exception as e:
                    logging.error(f"Erro ao processar linha para parcelamento: {e}")
                    resultado = {
                        "num_parcelamento": num_parcelamento if 'num_parcelamento' in locals() else "Desconhecido",
                        "nome_empresa": nome_empresa if 'nome_empresa' in locals() else "Desconhecido",
                        "sucesso": False,
                        "data_vencimento": None,
                        "motivo_falha": str(e)[:100]
                    }
                    pintar_planilha(caminho_planilha, resultado)
                    continue
            
            logging.info(f"Processamento concluído para {len(dados_filtrados)} parcelamentos")
        
        else:
            logging.error("Não foi possível carregar os dados da planilha-mãe")
    
    except Exception as e:
        logging.error(f"Erro no processo principal: {e}")
    
    finally:
        try:
            if 'driver' in locals() and driver is not None:
                try:
                    driver.quit()
                except Exception as e:
                    logging.warning(f"Erro ao encerrar WebDriver: {e}")
                finally:
                    driver = None
            logging.info("Navegador fechado")
        except Exception:
            pass

if __name__ == "__main__":
    try:
        logging.info("Iniciando script de automação de parcelamentos")
        main()
        logging.info("Script finalizado com sucesso")
    
    except Exception as e:
        logging.error(f"Erro fatal: {e}")
    
    print("\nProcessamento concluído!")
    sys.exit(0)
